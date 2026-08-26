"""Project Performance Monitoring — Excel data loader, seed generator and geo filters."""
from __future__ import annotations

import json
import time
from datetime import datetime
from typing import Any

from services.paths import BASE_DIR

# ────────────────────────────────────────────────────────────
# PROJECT PERFORMANCE MONITORING — Excel Data Loader
# ────────────────────────────────────────────────────────────
PERFORMANCE_XLSX_PATH = BASE_DIR.parent / "CHAK Monthly Project Performance Monitoring Template - Topline Indicators.xlsx"

_PROJECT_CACHE: dict[str, Any] | None = None
_PROJECT_CACHE_MTIME: float = 0

PROJECT_SHEET_MAP = {
    "jamii-tekelezi": "Jamii Tekelezi",
    "chap-stawisha": "CHAP Stawisha",
    "eye-health": "Eye Health - ACSP & GitLab",
    "eis": "EIS",
    "bftw-hss": "BFTW HSS",
    "bftw-rmncah": "BFTW RMNCAH",
    "pep": "PEP",
    "gf-mnch": "GF-MNCH",
    "impact": "IMPACT",
    "cdic-icare": "CDIC-iCARE",
}

BUDGET_LINE_NAMES = [
    "Personnel & HR",
    "Technical / Program Activities",
    "Equipment, Supplies & Commodities",
    "Training & Capacity Building",
    "Sub-awards / Partner Disbursements",
    "Monitoring, Evaluation & Learning",
    "Travel & Transport",
    "Administration & Overheads",
    "Other Direct Costs / Contingency",
]


def _parse_cell(cell: Any) -> float | str | None:
    """Parse a cell value: return float if numeric, stripped string if text, None if blank."""
    # Handle openpyxl Cell objects
    if hasattr(cell, "value"):
        value = cell.value
    else:
        value = cell
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, datetime):
        return value.strftime("%B %Y")
    s = str(value).strip()
    if not s:
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return s


def _rag_class(rag: str | None) -> str:
    """Normalize RAG string."""
    if not rag:
        return "N/A"
    r = str(rag).strip().lower()
    if r == "on track":
        return "On Track"
    if r == "watch":
        return "Watch"
    if r == "off track":
        return "Off Track"
    return str(rag).strip()


def load_project_performance_data(force_reload: bool = False) -> dict[str, Any]:
    """Load and cache the project performance Excel workbook. Returns a dict with
    'portfolio' (aggregate dashboard) and 'projects' (per-project detail)."""
    global _PROJECT_CACHE, _PROJECT_CACHE_MTIME

    xlsx = PERFORMANCE_XLSX_PATH
    if not xlsx.exists():
        return {"error": f"File not found: {xlsx.name}"}

    current_mtime = xlsx.stat().st_mtime
    if not force_reload and _PROJECT_CACHE is not None and current_mtime == _PROJECT_CACHE_MTIME:
        return _PROJECT_CACHE

    import openpyxl
    wb = openpyxl.load_workbook(xlsx, data_only=True)

    # ── Portfolio Dashboard Sheet ──
    portfolio_raw = _parse_portfolio_dashboard(wb["Portfolio Dashboard"])
    # ── Project Sheets ──
    projects: dict[str, Any] = {}
    for slug, sheet_name in PROJECT_SHEET_MAP.items():
        if sheet_name in wb.sheetnames:
            projects[slug] = _parse_project_sheet(wb[sheet_name], slug)

    wb.close()

    result = {
        "portfolio": portfolio_raw,
        "projects": projects,
        "loaded_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "sheet_count": len(PROJECT_SHEET_MAP),
    }
    _PROJECT_CACHE = result
    _PROJECT_CACHE_MTIME = current_mtime
    return result


def _parse_portfolio_dashboard(ws: Any) -> dict[str, Any]:
    """Parse the Portfolio Dashboard sheet."""
    rows_data: list[dict[str, Any]] = []
    # Data rows 6-15 (Excel rows); row 5 is header
    for excel_row in range(6, 16):
        proj = _parse_cell(ws.cell(row=excel_row, column=2))
        if not proj or proj == "PORTFOLIO TOTAL":
            break
        rows_data.append({
            "project": str(proj),
            "donor": str(_parse_cell(ws.cell(row=excel_row, column=3)) or ""),
            "total_annual_budget": _parse_cell(ws.cell(row=excel_row, column=4)) or 0,
            "cumulative_expenditure": _parse_cell(ws.cell(row=excel_row, column=5)) or 0,
            "budget_variance_pct": _parse_cell(ws.cell(row=excel_row, column=6)),
            "financial_rag": _rag_class(_parse_cell(ws.cell(row=excel_row, column=7))),
            "technical_rag": _rag_class(_parse_cell(ws.cell(row=excel_row, column=8))),
            "indicators_off_track": str(_parse_cell(ws.cell(row=excel_row, column=9)) or ""),
            "overall_rag": _rag_class(_parse_cell(ws.cell(row=excel_row, column=10))),
        })

    # Portfolio total (row 16)
    total_annual = _parse_cell(ws.cell(row=16, column=4)) or 0
    total_expenditure = _parse_cell(ws.cell(row=16, column=5)) or 0

    # CEO Snapshot (rows 19-20)
    on_track_count = str(_parse_cell(ws.cell(row=19, column=4)) or "")
    on_watch_count = str(_parse_cell(ws.cell(row=19, column=7)) or "")
    off_track_count = str(_parse_cell(ws.cell(row=20, column=4)) or "")
    budget_utilisation = _parse_cell(ws.cell(row=20, column=7))

    return {
        "projects": rows_data,
        "portfolio_total_annual_budget": total_annual,
        "portfolio_total_expenditure": total_expenditure,
        "ceo_snapshot": {
            "on_track": on_track_count,
            "on_watch": on_watch_count,
            "off_track": off_track_count,
            "budget_utilisation_pct": round(budget_utilisation * 100, 1) if isinstance(budget_utilisation, (int, float)) else None,
        },
    }


def _parse_project_sheet(ws: Any, slug: str) -> dict[str, Any]:
    """Parse an individual project sheet. Returns structured sections A, B, C."""
    # ── Header ──
    donor = str(_parse_cell(ws.cell(row=4, column=3)) or "")
    project_code = str(_parse_cell(ws.cell(row=5, column=3)) or "")
    reporting_month = ws.cell(row=4, column=7).value
    project_duration = _parse_cell(ws.cell(row=4, column=11))
    months_elapsed = _parse_cell(ws.cell(row=5, column=11))

    # Format reporting month
    if isinstance(reporting_month, datetime):
        reporting_month_str = reporting_month.strftime("%B %Y")
    else:
        reporting_month_str = str(reporting_month or "")

    # ── SECTION A: Financial Performance ──
    budget_lines: list[dict[str, Any]] = []
    for i, name in enumerate(BUDGET_LINE_NAMES):
        row = 9 + i
        budget_lines.append({
            "budget_line": name,
            "annual_budget": _parse_cell(ws.cell(row=row, column=3)) or 0,
            "planned_cumulative": _parse_cell(ws.cell(row=row, column=4)) or 0,
            "actual_cumulative": _parse_cell(ws.cell(row=row, column=5)) or 0,
            "variance_amount": _parse_cell(ws.cell(row=row, column=6)),
            "variance_pct": _parse_cell(ws.cell(row=row, column=7)),
            "current_month_expenditure": _parse_cell(ws.cell(row=row, column=8)) or 0,
            "avg_monthly_burn_rate": _parse_cell(ws.cell(row=row, column=9)) or 0,
            "projected_annual_expenditure": _parse_cell(ws.cell(row=row, column=10)) or 0,
            "months_remaining": _parse_cell(ws.cell(row=row, column=11)),
            "rag": _rag_class(_parse_cell(ws.cell(row=row, column=12))),
        })

    # Total row (row 18)
    total_budget_line = {
        "budget_line": "TOTAL PROJECT BUDGET",
        "annual_budget": _parse_cell(ws.cell(row=18, column=3)) or 0,
        "planned_cumulative": _parse_cell(ws.cell(row=18, column=4)) or 0,
        "actual_cumulative": _parse_cell(ws.cell(row=18, column=5)) or 0,
        "variance_amount": _parse_cell(ws.cell(row=18, column=6)),
        "variance_pct": _parse_cell(ws.cell(row=18, column=7)),
        "current_month_expenditure": _parse_cell(ws.cell(row=18, column=8)) or 0,
        "avg_monthly_burn_rate": _parse_cell(ws.cell(row=18, column=9)) or 0,
        "projected_annual_expenditure": _parse_cell(ws.cell(row=18, column=10)) or 0,
        "months_remaining": _parse_cell(ws.cell(row=18, column=11)),
        "rag": _rag_class(_parse_cell(ws.cell(row=18, column=12))),
    }

    # ── SECTION B: Technical / Donor Indicators ──
    indicators: list[dict[str, Any]] = []
    for row in range(23, 31):
        ind_name = _parse_cell(ws.cell(row=row, column=2))
        if not ind_name:
            continue
        indicators.append({
            "indicator": str(ind_name),
            "definition": str(_parse_cell(ws.cell(row=row, column=3)) or ""),
            "annual_target": _parse_cell(ws.cell(row=row, column=4)) or 0,
            "planned_cumulative": _parse_cell(ws.cell(row=row, column=5)) or 0,
            "actual_cumulative": _parse_cell(ws.cell(row=row, column=6)) or 0,
            "achievement_pct": _parse_cell(ws.cell(row=row, column=7)),
            "rag": _rag_class(_parse_cell(ws.cell(row=row, column=12))),
        })

    # ── SECTION C: Overall Project Health ──
    financial_rag = _rag_class(_parse_cell(ws.cell(row=34, column=5)))
    off_track_budget_lines = str(_parse_cell(ws.cell(row=34, column=10)) or "")
    technical_rag = _rag_class(_parse_cell(ws.cell(row=35, column=5)))
    off_track_indicators = str(_parse_cell(ws.cell(row=35, column=10)) or "")
    overall_rag = _rag_class(_parse_cell(ws.cell(row=36, column=5)))

    # ── Programme Manager's Narrative (rows 38-39) ──
    narrative_key_achievements = str(_parse_cell(ws.cell(row=38, column=2)) or "")
    narrative_text = str(_parse_cell(ws.cell(row=39, column=2)) or "")

    # Compute financial variance pct from total budget line
    total_annual = total_budget_line["annual_budget"] or 0
    total_actual = total_budget_line["actual_cumulative"] or 0
    total_planned = total_budget_line["planned_cumulative"] or 0
    total_variance_pct = total_budget_line["variance_pct"]

    return {
        "slug": slug,
        "donor": donor,
        "project_code": project_code,
        "reporting_month": reporting_month_str,
        "project_duration_months": project_duration,
        "months_elapsed": months_elapsed,
        "section_a": {
            "budget_lines": budget_lines,
            "total": total_budget_line,
            "total_annual_budget": total_annual,
            "total_cumulative_expenditure": total_actual,
            "total_planned_cumulative": total_planned,
            "total_variance_pct": total_variance_pct,
        },
        "section_b": {
            "indicators": indicators,
        },
        "section_c": {
            "financial_rag": financial_rag,
            "off_track_budget_lines": off_track_budget_lines,
            "technical_rag": technical_rag,
            "off_track_indicators": off_track_indicators,
            "overall_rag": overall_rag,
        },
        "narrative": {
            "key_achievements": narrative_key_achievements,
            "narrative_text": narrative_text,
        },
    }


# ── Cover & Instructions Sheet ────────────────────────────────
def _parse_cover_sheet(ws: Any) -> dict[str, Any]:
    """Parse the Cover & Instructions sheet."""
    return {
        "title": str(_parse_cell(ws.cell(row=2, column=2)) or ""),
        "instructions": [
            str(_parse_cell(ws.cell(row=r, column=2)) or "")
            for r in range(7, 20)
            if _parse_cell(ws.cell(row=r, column=2))
        ],
        "color_key": {
            "blue_input": "Input cell — edit this",
            "black_formula": "Formula — do not edit",
            "green_link": "Link pulling from another tab",
        },
        "rag_legend": ["On Track", "Watch", "Off Track"],
    }


# ────────────────────────────────────────────────────────────
# SEED DATA GENERATOR — geographic breakdowns
# ────────────────────────────────────────────────────────────
_SEED_CACHE: dict | None = None
_SEED_CACHE_PATH = BASE_DIR / "data" / "performance_seed.json"

# Realistic Programme Manager's Narratives for each project
PROJECT_NARRATIVES = {
    "jamii-tekelezi": {
        "key_achievements": "Key achievements during this reporting period: Successfully reached 85% of enrolled beneficiaries with comprehensive HIV prevention services. Conducted 12 community outreach campaigns across Meru and Tharaka Nithi counties. Trained 45 community health volunteers on client-centered HIV testing approaches.",
        "narrative_text": "Programme Manager's Narrative: Jamii Tekelezi continues to show strong community engagement, though financial performance requires attention. The over-expenditure on Personnel (22% above plan) is due to the recruitment of additional outreach workers to cover hard-to-reach areas. Technical indicators are largely on track with 7 of 8 achieving targets. The project has strengthened linkages between community-based testing and facility-based ART initiation. Recommendation: Review personnel budget allocation for Q3 to align with actual staffing needs."
    },
    "chap-stawisha": {
        "key_achievements": "Key achievements: Launched the CHAP Stawisha program across 12 new facilities in Nyandarua County. Completed baseline assessment for 2,400 OVC caregivers. Established 8 community support groups for adolescent girls and young women (AGYW).",
        "narrative_text": "Programme Manager's Narrative: CHAP Stawisha is in early implementation phase. Financial data shows zero expenditure as the project is awaiting initial disbursement from CDC. Technical indicators show 50% achievement on Tracer Indicator 1 (baseline assessments completed). Other indicators are still at planning stage. The project team is fully recruited and orientation has been completed. Recommendation: Fast-track initial fund disbursement to enable activity implementation in Q3."
    },
    "eye-health": {
        "key_achievements": "Key achievements: Conducted comprehensive eye camps in 8 sub-counties reaching 3,200 patients. Distributed 1,850 reading glasses and performed 240 cataract surgeries. Trained 32 primary health care workers on basic eye examination protocols.",
        "narrative_text": "Programme Manager's Narrative: Eye Health project is performing well across all indicators. Financial burn rate is within acceptable range (-10% variance) with strong expenditure controls. Technical indicators all above 90% achievement. The project has established strong referral networks between community health workers and eye units at sub-county hospitals. Procurement of additional ophthalmic equipment is in progress. Recommendation: Consider scaling the community eye camp model to cover remaining underserved sub-counties."
    },
    "eis": {
        "key_achievements": "Key achievements: Completed integrated service delivery assessments in 15 facilities across 4 counties. Developed comprehensive EIS implementation roadmap. Trained 60 healthcare workers on integrated service delivery protocols.",
        "narrative_text": "Programme Manager's Narrative: EIS is facing significant challenges with budget over-expenditure (25% above plan). This is primarily due to higher-than-anticipated costs for field assessments and training logistics. Technical indicators are also below target (average 70% achievement). The project scope is broad and requires more robust planning for resource allocation. An internal review is scheduled for next month to revise the activity budget and work plan. Recommendation: Conduct mid-project review to realign budget and activities with available resources."
    },
    "bftw-hss": {
        "key_achievements": "Key achievements: Strengthened health systems in 20 faith-based health facilities. Completed infrastructure assessments and developed prioritized improvement plans. Trained 55 facility managers on financial management and governance.",
        "narrative_text": "Programme Manager's Narrative: BFTW HSS is progressing satisfactorily. Financial performance is on track with -3% variance (slightly under budget). Technical indicators are at Watch status (88% achievement) — close to target but needing focused attention. Facility infrastructure upgrades are proceeding according to schedule. The project has successfully integrated financial management training with hands-on mentoring. Recommendation: Intensify technical support to facilities that are lagging in indicator achievement."
    },
    "bftw-rmncah": {
        "key_achievements": "Key achievements: Supported RMNCAH services in 18 CHAK member facilities. Conducted maternal and newborn health mentorship sessions reaching 120 healthcare workers. Distributed essential RMNCAH supplies and equipment to 12 high-volume facilities.",
        "narrative_text": "Programme Manager's Narrative: BFTW RMNCAH is on track both financially and technically. Budget variance is within the +5% tolerance and all indicators are above 92% achievement. The project has contributed to improved maternal health outcomes in supported facilities. Skilled birth attendance has increased by 15% compared to baseline. The mentorship model continues to show positive impact on quality of care. Recommendation: Document best practices for replication across the CHAK network."
    },
    "pep": {
        "key_achievements": "Key achievements: Launched professional education programs for health professionals. Enrolled 85 students in partnership programs. Established clinical placement agreements with 10 training hospitals.",
        "narrative_text": "Programme Manager's Narrative: PEP (Partnership for Education of Health Professionals) is in early implementation (4 months elapsed of 18-month project). The Novo Nordisk-funded project is establishing partnerships with training institutions and hospitals. Financial variance is -22% due to the initial setup phase — recruitment of faculty and procurement of training materials are in progress. Technical indicators are on track but based on planned vs. actual enrollment. Recommendation: Accelerate procurement processes to align expenditure with the implementation timeline."
    },
    "gf-mnch": {
        "key_achievements": "Key achievements: Implemented Gates Foundation MNCH program across 25 high-burden facilities. Achieved 95% coverage for intermittent preventive treatment in pregnancy (IPTp). Distributed 8,500 long-lasting insecticidal nets to pregnant women and children under five.",
        "narrative_text": "Programme Manager's Narrative: GF-MNCH is performing excellently with both financial (1% variance) and technical indicators on track. The project has achieved high coverage rates for key MNCH interventions. Community health workers have been instrumental in reaching pregnant women with timely services. The project's data quality improvement initiatives have strengthened HMIS reporting at facility level. Recommendation: Explore opportunities for integration with other maternal health programs for sustainable impact."
    },
    "impact": {
        "key_achievements": "Key achievements: Project implementation at full scale across all target counties. Reached 15,000 direct beneficiaries with comprehensive health interventions. Established strong community accountability mechanisms.",
        "narrative_text": "Programme Manager's Narrative: IMPACT project (funded by action medeor) is significantly over-budget (30% variance) with all budget lines off track. The over-expenditure is due to expanded geographic coverage requested by the donor, coupled with rising operational costs. Technical indicators are similarly off track across all 8 indicators (average 56% achievement). A comprehensive restructuring of the project implementation plan is required. Recommendation: Engage with action medeor for a no-cost extension and budget realignment. Consider reducing scope to match available resources."
    },
    "cdic-icare": {
        "key_achievements": "Key achievements: Scaled up iCARE program to reach 22 health facilities. Trained 110 healthcare workers on integrated chronic disease management. Established 8 chronic disease support groups at facility level.",
        "narrative_text": "Programme Manager's Narrative: CDIC-iCARE (CDC/iCARE funded) is on track financially (8% variance) but technical indicators are at Watch status (average 88% achievement). The chronic disease management program has been well-received in target facilities. The project has successfully integrated HIV and NCD services in 18 facilities. Challenges remain in consistent supply of essential NCD medicines. Recommendation: Strengthen supply chain coordination with county health teams to ensure consistent availability of NCD commodities."
    },
}


def _load_geography_hierarchy() -> dict:
    """Load county/subcounty/facility hierarchy from jamii_tekelezi_filters.csv."""
    jt_path = BASE_DIR / "data" / "jamii_tekelezi_filters.csv"
    if not jt_path.exists():
        return {"counties": [], "subcounties": [], "facilities": [], "hierarchy": {}}
    import pandas as _pd
    df = _pd.read_csv(jt_path)
    hierarchy = {}
    for _, row in df.iterrows():
        c_name = str(row.get("county_name", "")).strip()
        sc_name = str(row.get("subcounty_name", "")).strip()
        f_name = str(row.get("facility_name", "")).strip()
        f_id = str(row.get("facility_id", "")).strip()
        if not c_name:
            continue
        if c_name not in hierarchy:
            hierarchy[c_name] = {"subcounties": {}}
        if sc_name not in hierarchy[c_name]["subcounties"]:
            hierarchy[c_name]["subcounties"][sc_name] = []
        if f_name:
            hierarchy[c_name]["subcounties"][sc_name].append({"id": f_id, "name": f_name})
    return {
        "counties": sorted(hierarchy.keys()),
        "subcounties": {c: sorted(h["subcounties"].keys()) for c, h in hierarchy.items()},
        "facilities": {f"{c}/{sc}": sorted(
            [{"id": fi["id"], "name": fi["name"]} for fi in facilities],
            key=lambda x: x["name"]
        ) for c, h in hierarchy.items() for sc, facilities in h["subcounties"].items()},
        "hierarchy": hierarchy,
    }


def _generate_performance_seed_data(force_reload: bool = False) -> dict:
    """Generate seed data for geographic breakdowns of project performance.
    Takes the Excel portfolio data and distributes it across counties/subcounties/facilities."""
    global _SEED_CACHE

    # Check cache
    if not force_reload and _SEED_CACHE is not None:
        return _SEED_CACHE
    if not force_reload and _SEED_CACHE_PATH.exists():
        try:
            with open(_SEED_CACHE_PATH) as f:
                _SEED_CACHE = json.load(f)
                return _SEED_CACHE
        except (json.JSONDecodeError, OSError):
            pass

    # Load base project data
    base_data = load_project_performance_data()
    if "error" in base_data:
        return base_data

    # Load geography
    geo = _load_geography_hierarchy()
    counties = geo["counties"]
    hierarchy = geo["hierarchy"]
    if not counties:
        return {"error": "No geography data found", "seed": False}

    import random
    random.seed(42)  # reproducible

    # Calculate facility count per county as weight for budget distribution
    county_facility_counts = {}
    for c_name, c_data in hierarchy.items():
        count = sum(len(scls) for scls in c_data["subcounties"].values())
        county_facility_counts[c_name] = max(count, 1)

    total_facilities = sum(county_facility_counts.values())

    # Build seeded project data
    seeded_projects = {}
    project_geo_data = {}

    for slug, project in base_data.get("projects", {}).items():
        sec_a = project.get("section_a", {})
        total_budget = sec_a.get("total_annual_budget", 0) or 0
        total_expenditure = sec_a.get("total_cumulative_expenditure", 0) or 0
        donor = project.get("donor", "")
        narrative = PROJECT_NARRATIVES.get(slug, {
            "key_achievements": "Key achievements narrative not yet submitted for this reporting period.",
            "narrative_text": "Programme Manager's Narrative: Detailed narrative report is pending submission from the project team. Will be updated in the next reporting cycle."
        })

        # Distribute budget across counties proportionally by facility count
        county_data = {}
        for c_name in counties:
            fac_count = county_facility_counts.get(c_name, 1)
            weight = fac_count / total_facilities

            county_budget = round(total_budget * weight, 0)
            # Expenditure varies slightly from proportional
            variance_factor = 1 + random.uniform(-0.15, 0.15)
            county_expenditure = round(total_expenditure * weight * variance_factor, 0)
            county_planned = round((total_budget * weight) * (project.get("months_elapsed", 6) or 6) / max(project.get("project_duration_months", 12) or 12, 1), 0)

            # Sub-county breakdown
            subcounties_data = {}
            c_hierarchy = hierarchy.get(c_name, {"subcounties": {}})
            sc_list = sorted(c_hierarchy["subcounties"].keys())
            sc_fac_counts = {}
            for sc_name, facilities in c_hierarchy["subcounties"].items():
                sc_fac_counts[sc_name] = len(facilities)
            sc_total_fac = max(sum(sc_fac_counts.values()), 1)

            for sc_name in sc_list:
                sc_weight = sc_fac_counts.get(sc_name, 1) / sc_total_fac
                sc_budget = round(county_budget * sc_weight, 0)
                sc_expenditure = round(county_expenditure * sc_weight * (1 + random.uniform(-0.1, 0.1)), 0)
                sc_planned = round(county_planned * sc_weight, 0)

                # Facility-level breakdown
                facilities_list = c_hierarchy["subcounties"].get(sc_name, [])
                facilities_data = []
                for fac in facilities_list:
                    if sc_total_fac > 0:
                        fac_weight = 1.0 / max(len(facilities_list), 1)
                    else:
                        fac_weight = 0
                    fac_budget = round(sc_budget * fac_weight, 0)
                    fac_expenditure = round(sc_expenditure * fac_weight * (1 + random.uniform(-0.05, 0.05)), 0)
                    facilities_data.append({
                        "facility_id": fac["id"],
                        "facility_name": fac["name"],
                        "allocated_budget": fac_budget,
                        "actual_expenditure": fac_expenditure,
                        "beneficiaries_served": random.randint(50, 2000),
                    })

                subcounties_data[sc_name] = {
                    "allocated_budget": sc_budget,
                    "actual_expenditure": sc_expenditure,
                    "planned_expenditure": sc_planned,
                    "facility_count": len(facilities_list),
                    "beneficiaries_served": random.randint(500, 15000),
                    "facilities": facilities_data,
                }

            county_data[c_name] = {
                "allocated_budget": county_budget,
                "actual_expenditure": county_expenditure,
                "planned_expenditure": county_planned,
                "facility_count": fac_count,
                "beneficiaries_served": random.randint(2000, 50000),
                "subcounties": subcounties_data,
            }

        seeded_projects[slug] = {
            "donor": donor,
            "counties": county_data,
        }
        project_geo_data[slug] = {
            "total_budget": total_budget,
            "total_expenditure": total_expenditure,
            "narrative": narrative,
        }

    result = {
        "seed": True,
        "counties": counties,
        "metadata": {
            "generated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
            "total_counties": len(counties),
            "total_facilities": total_facilities,
            "seed_version": "1.0",
        },
        "projects": seeded_projects,
        "project_metadata": project_geo_data,
    }

    # Cache to file
    try:
        _SEED_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
        with open(_SEED_CACHE_PATH, "w") as f:
            json.dump(result, f, indent=2, default=str)
    except OSError:
        pass

    _SEED_CACHE = result
    return result


def _filter_performance_by_geography(
    county: str | None = None,
    subcounty: str | None = None,
    facility: str | None = None,
    project_slug: str | None = None,
) -> dict:
    """Filter project performance data by geographic dimensions.
    Returns aggregated data based on the selected geography."""
    seed = _generate_performance_seed_data()
    if "error" in seed:
        return seed

    base = load_project_performance_data()
    if "error" in base:
        return base

    result = {
        "portfolio": base.get("portfolio", {}),
        "projects": {},
        "filter_applied": {
            "county": county or "all",
            "subcounty": subcounty or "all",
            "facility": facility or "all",
        },
        "seeded": True,
    }

    seed_projects = seed.get("projects", {})
    base_projects = base.get("projects", {})

    target_slugs = [project_slug] if project_slug else list(base_projects.keys())

    for slug in target_slugs:
        if slug not in base_projects:
            continue
        base_proj = base_projects[slug]
        seed_proj = seed_projects.get(slug, {})

        # Start with base project data
        filtered = dict(base_proj)
        geo_data = {"counties": {}}

        for c_name, c_data in seed_proj.get("counties", {}).items():
            # Apply county filter
            if county and county.lower() != c_name.lower().replace(" ", "_").replace("-", "_"):
                if county.lower() != c_name.lower():
                    continue

            sc_data_filtered = {}
            for sc_name, sc_data in c_data.get("subcounties", {}).items():
                # Apply subcounty filter
                if subcounty and subcounty.lower() != sc_name.lower().replace(" ", "_").replace("-", "_"):
                    if subcounty.lower() != sc_name.lower():
                        continue

                fac_data_filtered = []
                for fac in sc_data.get("facilities", []):
                    # Apply facility filter
                    if facility and facility.lower() != fac["facility_name"].lower().replace(" ", "_").replace("-", "_"):
                        if facility.lower() != fac["facility_name"].lower():
                            continue
                    fac_data_filtered.append(fac)

                if not facility or fac_data_filtered:
                    sc_entry = dict(sc_data)
                    sc_entry["facilities"] = fac_data_filtered
                    sc_data_filtered[sc_name] = sc_entry

            if not subcounty or sc_data_filtered:
                c_entry = dict(c_data)
                c_entry["subcounties"] = sc_data_filtered
                geo_data["counties"][c_name] = c_entry

        filtered["geo_breakdown"] = geo_data
        filtered["narrative"] = seed.get("project_metadata", {}).get(slug, {}).get("narrative", base_proj.get("narrative", {}))
        result["projects"][slug] = filtered

    return result
