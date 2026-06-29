from __future__ import annotations

import gzip
import html
import json
import os
import re
import sqlite3
import threading
import time
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any

import xml.etree.ElementTree as ET

import pandas as pd
from dotenv import load_dotenv
from flask import Flask, jsonify, request, send_from_directory

try:
    import google.generativeai as genai
except ImportError:  # pragma: no cover - handled at runtime if package is missing
    genai = None


BASE_DIR = Path(__file__).resolve().parent
CSV_PATH = BASE_DIR / "golden_executive_record.csv"
GUIDE_XLSX_PATH = BASE_DIR / "Copy of DATIM DATA ENTRY GUIDE FY26 Q2.xlsx"
TABLE_NAME = "clinics"

# ── Superpower module for DHIS2 live queries ──────────────────────
SUPERPOWER_DIR = BASE_DIR.parent  # ai_translator.py is in the repo root
SUPERPOWER_ENV_PATH = BASE_DIR / ".env"
if SUPERPOWER_ENV_PATH.exists():
    _sp_vars = {}
    with open(SUPERPOWER_ENV_PATH) as _f:
        for _line in _f:
            _line = _line.strip()
            if _line and "=" in _line and not _line.startswith("#"):
                _k, _v = _line.split("=", 1)
                _sp_vars[_k.strip()] = _v.strip()
    # Override DHIS credentials so superpower uses the CHAK server
    for _k, _v in _sp_vars.items():
        if _k.startswith("DHIS_") or _k.startswith("GEMINI_API_KEY"):
            os.environ[_k] = _v

import sys
sys.path.insert(0, str(SUPERPOWER_DIR))
try:
    from ai_translator import generate_dhis2_url as _superpower_generate_url
    from ai_translator import fetch_query_result as _superpower_fetch_result
    from ai_translator import load_dictionaries as _superpower_load_dict
    HAS_SUPERPOWER = True
except Exception:
    HAS_SUPERPOWER = False
    _superpower_generate_url = None
    _superpower_fetch_result = None
    _superpower_load_dict = None
MAX_RESULT_ROWS = 100

load_dotenv()


def _period_sort_key(period_name):
    """Sort month names chronologically: 'June 2025' < 'September 2025' < 'April 2026'.
    Handles period codes like '202509' as fallback."""
    MONTHS = {
        "January": 1, "February": 2, "March": 3, "April": 4,
        "May": 5, "June": 6, "July": 7, "August": 8,
        "September": 9, "October": 10, "November": 11, "December": 12,
    }
    s = str(period_name)
    parts = s.split()
    if len(parts) >= 2 and parts[0] in MONTHS:
        return (int(parts[1]), MONTHS[parts[0]])
    # Fallback: assume period code like "202509"
    if len(s) == 6 and s.isdigit():
        return (int(s[:4]), int(s[4:]))
    return (0, 0, s)  # put unrecognized at end


# ── JT filter → DHIS2 OU ID resolution ──
_JT_OU_CACHE = None  # {facility_name: facility_id, subcounty_name: [facility_id, ...]}


def _load_jt_ou_map():
    """Load Jamii Tekelezi filter CSV into OU lookup maps."""
    global _JT_OU_CACHE
    if _JT_OU_CACHE is not None:
        return _JT_OU_CACHE
    jt_path = BASE_DIR / "data" / "jamii_tekelezi_filters.csv"
    if not jt_path.exists():
        _JT_OU_CACHE = {}
        return _JT_OU_CACHE
    import pandas as _pd
    jt_df = _pd.read_csv(jt_path)
    # facility name → OU ID
    name_to_id = {}
    subcounty_to_ids = {}
    for _, row in jt_df.iterrows():
        fid = str(row.get("facility_id", "")).strip()
        fname = str(row.get("facility_name", "")).strip()
        sc = str(row.get("subcounty_name", "")).strip()
        if fid and fname:
            name_to_id[fname] = fid
        if sc and fid:
            subcounty_to_ids.setdefault(sc, []).append(fid)
    # deduplicate subcounty lists
    for sc in subcounty_to_ids:
        subcounty_to_ids[sc] = list(dict.fromkeys(subcounty_to_ids[sc]))
    _JT_OU_CACHE = {"name_to_id": name_to_id, "subcounty_to_ids": subcounty_to_ids}
    return _JT_OU_CACHE


def _resolve_ou_ids(county, subcounty=None, facility=None):
    """Resolve facility/subcounty filters to DHIS2 OU IDs.
    Returns (ou_id_or_list, is_multi) where is_multi means we have multiple OUs.
    """
    JT_COUNTY_IDS = {
        "Meru County": "Y52XNJ50hYb",
        "Embu County": "PFu8alU2KWG",
        "Nyandarua County": "mYZacFNIB3h",
        "Tharaka Nithi County": "T4urHM47nlm",
    }
    default_ou = JT_COUNTY_IDS.get(county, "Y52XNJ50hYb")

    if not facility and not subcounty:
        return default_ou, False

    jt_map = _load_jt_ou_map()
    if not jt_map:
        return default_ou, False

    # Facility takes precedence
    if facility and facility != "all":
        fid = jt_map["name_to_id"].get(facility)
        if fid:
            return fid, False
        # facility not found in JT, fall through to subcounty or county

    # Subcounty → list of facility IDs
    if subcounty and subcounty != "all":
        ids = jt_map["subcounty_to_ids"].get(subcounty, [])
        if ids:
            return ids, True

    return default_ou, False


def create_app() -> Flask:
    app = Flask(__name__)
    app.config["JSON_SORT_KEYS"] = False

    database = initialize_database(CSV_PATH)
    dataframe = load_source_dataframe(CSV_PATH)
    schema_sql = pd.io.sql.get_schema(dataframe, TABLE_NAME, con=database)

    app.config["DATABASE_CONNECTION"] = database
    app.config["DATABASE_LOCK"] = threading.Lock()
    app.config["DATAFRAME"] = dataframe
    app.config["SCHEMA_SQL"] = schema_sql
    app.config["GEMINI_PROVIDERS"] = build_gemini_providers()
    app.config["GEMINI_ROUTER_STATE"] = {"cursor": 0, "cooldowns": {}}
    app.config["AI_RESPONSE_CACHE"] = {}
    app.config["AI_RESPONSE_CACHE_LOCK"] = threading.Lock()
    app.config["CATALOG"] = build_canonical_catalog(dataframe)
    app.config["SOURCE_DATASETS"] = {}
    app.config["LOCATION_HIERARCHY"] = load_datim_location_hierarchy(GUIDE_XLSX_PATH)
    # CSV loader paused — kept for future reference:
    # csv_hierarchy_path = BASE_DIR / "location_hierarchy.csv"
    # app.config["LOCATION_HIERARCHY"] = load_location_hierarchy_from_csv(csv_hierarchy_path)
    # if not app.config["LOCATION_HIERARCHY"].get("counties"):
    #     app.config["LOCATION_HIERARCHY"] = load_datim_location_hierarchy(GUIDE_XLSX_PATH)
    app.config["DATIM_HIV_TREATMENT_SECTIONS"] = load_datim_hiv_treatment_sections(GUIDE_XLSX_PATH)
    # track additional CSVs loaded from data/dhis
    app.config["LOADED_CSVS"] = set()

    # Purge any stale fallback/local cached responses that may exist from previous runs.
    try:
        with app.config["AI_RESPONSE_CACHE_LOCK"]:
            keys_to_remove = [k for k, v in (app.config.get("AI_RESPONSE_CACHE") or {}).items() if isinstance(v, dict) and v.get("source") in ("fallback", "no_gemini", "local")]
            for k in keys_to_remove:
                app.config["AI_RESPONSE_CACHE"].pop(k, None)
            # If any entries remain, ensure the cache is empty on fresh startup
            if app.config.get("AI_RESPONSE_CACHE"):
                app.config["AI_RESPONSE_CACHE"].clear()
    except Exception:
        # don't fail startup for caching cleanup issues
        pass

    def scan_and_load_additional_csvs():
        """Scan data/dhis/raw and data/dhis/processed for CSV files and merge them into the main dataframe.

        This function is idempotent and will skip files already loaded. It attempts to transform
        DHIS analytics rows (dx,pe,ou,value) into friendly columns (Indicator, Month, Facility,
        Total_Visits / Monthly_Expenditure_USD) before merging so the UI filters can pick them up.
        """
        raw_dir = BASE_DIR / "data" / "dhis" / "raw"
        proc_dir = BASE_DIR / "data" / "dhis" / "processed"
        candidates = []
        for folder in (raw_dir, proc_dir):
            try:
                for p in folder.glob("*.csv"):
                    candidates.append(p)
            except Exception:
                continue

        new_files = [p for p in candidates if str(p) not in app.config["LOADED_CSVS"]]
        if not new_files:
            return []

        # load metadata maps if available
        meta_dir = BASE_DIR / "data" / "dhis" / "meta"
        data_elements_map = {}
        org_units_map = {}
        try:
            de_path = meta_dir / "data_elements.csv"
            if de_path.exists():
                df_de = pd.read_csv(de_path)
                if "id" in df_de.columns and "name" in df_de.columns:
                    data_elements_map = dict(zip(df_de["id"].astype(str), df_de["name"]))
        except Exception:
            pass
        try:
            ind_path = meta_dir / "indicators.csv"
            if ind_path.exists():
                df_ind = pd.read_csv(ind_path)
                if "id" in df_ind.columns and "name" in df_ind.columns:
                    for indicator_id, indicator_name in zip(df_ind["id"].astype(str), df_ind["name"]):
                        data_elements_map.setdefault(indicator_id, indicator_name)
        except Exception:
            pass
        try:
            ou_path = meta_dir / "organisation_units.csv"
            if ou_path.exists():
                df_ou = pd.read_csv(ou_path)
                if "id" in df_ou.columns:
                    org_units_map = {
                        clean_text(row.get("id")): row.to_dict()
                        for _, row in df_ou.iterrows()
                        if clean_text(row.get("id"))
                    }
        except Exception:
            pass

        app.config["ORG_UNITS_LOOKUP"] = org_units_map

        hospital_map: dict[str, dict[str, Any]] = {}
        try:
            hospitals_path = proc_dir / "hospitals.csv"
            if hospitals_path.exists():
                df_hospitals = pd.read_csv(hospitals_path)
                app.config["HOSPITALS_TABLE"] = df_hospitals.copy()
                required_columns = {"hospital_id", "hospital_name", "path", "level"}
                if required_columns.issubset(df_hospitals.columns):
                    hospital_map = {
                        clean_text(row.get("hospital_id")): {
                            "hospital_name": clean_text(row.get("hospital_name")),
                            "path": clean_text(row.get("path")),
                            "level": row.get("level"),
                        }
                        for _, row in df_hospitals.iterrows()
                        if clean_text(row.get("hospital_id"))
                    }
        except Exception:
            app.logger.exception("Failed to preload hospital metadata")

        app.config["HOSPITALS_LOOKUP"] = hospital_map

        frames = []
        source_datasets: dict[str, pd.DataFrame] = {}
        for p in new_files:
            try:
                df = pd.read_csv(p)
                if df.empty:
                    app.logger.debug("Skipping empty CSV: %s", p)
                    app.config["LOADED_CSVS"].add(str(p))
                    continue

                lower_name = p.name.lower()
                if lower_name == "hospitals.csv":
                    app.config["LOADED_CSVS"].add(str(p))
                    app.logger.info("Loaded hospital lookup table: %s (%d rows)", p.name, len(df))
                    continue

                if lower_name == "dhis_combined.csv":
                    app.config["LOADED_CSVS"].add(str(p))
                    app.logger.info("Skipped pre-combined DHIS CSV: %s", p.name)
                    continue

                # detect DHIS analytics table format: dx,pe,ou,value
                cols = [c.lower() for c in df.columns]
                if set(cols) >= {"dx", "pe", "ou", "value"}:
                    transformed = normalize_dhis_analytics_frame(
                        df,
                        p.name,
                        data_elements_map,
                        org_units_map,
                        hospital_map,
                    )
                    frames.append(transformed)
                    source_datasets[p.name] = transformed
                    app.config["LOADED_CSVS"].add(str(p))
                    app.logger.info("Transformed DHIS analytics CSV: %s -> %d rows", p.name, len(transformed))
                    continue

                # otherwise append as-is but ensure no duplicate column names
                app.config["LOADED_CSVS"].add(str(p))
                app.logger.info("Loaded non-analytics CSV without merging: %s (%d rows)", p.name, len(df))
            except Exception as exc:
                app.logger.warning("Failed to load CSV %s: %s", p, exc)

        if not frames:
            return new_files

        try:
            # Always keep the original CSV-backed dataframe and append any newly loaded DHIS frames.
            base_frames = [app.config["DATAFRAME"]] + frames
            combined = pd.concat(base_frames, ignore_index=True, sort=False)
            # normalize column order by reindexing
            combined = combined.loc[:, ~combined.columns.duplicated()]
            # Normalize Month column to consistent YYYYMM string where possible to avoid mixed-type sorting errors
            if "Month" in combined.columns:
                s = combined["Month"].fillna("").astype(str).str.strip()
                # remove non-digits to handle formats like '2025-12' or '2025/12'
                s_clean = s.str.replace(r"[^0-9]", "", regex=True)
                # keep only first 6 digits if present
                s_clean = s_clean.str.slice(0, 6)
                mask6 = s_clean.str.match(r"^\d{6}$")
                # where we have a 6-digit YYYYMM, use it; otherwise fall back to original string (trimmed)
                combined["Month"] = s_clean.where(mask6, s)
                # ensure uniform string dtype
                combined["Month"] = combined["Month"].astype(str)

            if "Facility" in combined.columns and hospital_map:
                facility_context = combined["Facility"].astype(str).map(
                    lambda unit_id: resolve_unit_context(unit_id, org_units_map, hospital_map)
                )
                combined["Facility"] = facility_context.map(lambda context: context["Facility"])

                county_values = facility_context.map(lambda context: context["County"])
                subcounty_values = facility_context.map(lambda context: context["SubCounty"])

                if "County" not in combined.columns:
                    combined["County"] = county_values
                else:
                    combined["County"] = combined["County"].fillna("").astype(str)
                    combined["County"] = combined["County"].where(combined["County"].str.strip() != "", county_values)

                if "SubCounty" not in combined.columns:
                    combined["SubCounty"] = subcounty_values
                else:
                    combined["SubCounty"] = combined["SubCounty"].fillna("").astype(str)
                    combined["SubCounty"] = combined["SubCounty"].where(combined["SubCounty"].str.strip() != "", subcounty_values)

            app.config["DATAFRAME"] = combined
            app.config["SOURCE_DATASETS"] = source_datasets
            # update sqlite table
            with app.config["DATABASE_LOCK"]:
                conn = app.config["DATABASE_CONNECTION"]
                combined.to_sql(TABLE_NAME, conn, index=False, if_exists="replace")
            # refresh schema and registries
            app.config["SCHEMA_SQL"] = pd.io.sql.get_schema(combined, TABLE_NAME, con=app.config["DATABASE_CONNECTION"]) if not combined.empty else app.config.get("SCHEMA_SQL")
            app.config["CATALOG"] = build_canonical_catalog(combined)
        except Exception as exc:
            app.logger.exception("Failed to merge additional CSVs: %s", exc)

        return new_files

    def csv_watcher_loop(interval_seconds: int = 20):
        # simple polling loop that runs in a daemon thread
        while True:
            try:
                added = scan_and_load_additional_csvs()
                if added:
                    app.logger.info("CSV watcher loaded %d new files", len(added))
                time.sleep(interval_seconds)
            except Exception:
                app.logger.exception("CSV watcher encountered an error")
                time.sleep(interval_seconds)

    # load once at startup so the first request sees the enriched data
    scan_and_load_additional_csvs()

    # start watcher thread
    watcher = threading.Thread(target=csv_watcher_loop, args=(20,), daemon=True)
    watcher.start()

    @app.after_request
    def compress_response(response):
        if response.direct_passthrough:
            return response

        accept_encoding = request.headers.get("Accept-Encoding", "")
        if "gzip" not in accept_encoding.lower():
            return response

        content_type = response.headers.get("Content-Type", "")
        if not any(
            token in content_type
            for token in ("application/json", "text/html", "text/javascript", "application/javascript")
        ):
            return response

        payload = response.get_data()
        if len(payload) < 1024 or response.headers.get("Content-Encoding"):
            return response

        compressed = gzip.compress(payload)
        response.set_data(compressed)
        response.headers["Content-Encoding"] = "gzip"
        response.headers["Content-Length"] = str(len(compressed))
        response.headers.add("Vary", "Accept-Encoding")
        return response

    @app.get("/")
    def index() -> object:
        index_path = BASE_DIR / "index.html"
        if index_path.exists():
            return send_from_directory(BASE_DIR, "index.html")
        return jsonify(
            {
                "message": "Executive Intelligence Dashboard backend is running.",
                "endpoints": ["/api/dashboard-data", "/api/chat"],
            }
        )

    @app.get("/api/health")
    def health() -> object:
        return jsonify(json_safe({"status": "ok"}))

    @app.get("/api/dashboard-data")
    def dashboard_data() -> object:
        return jsonify(
            json_safe(
                {
                    "table": TABLE_NAME,
                    "row_count": int(len(app.config["DATAFRAME"])),
                    "columns": list(app.config["DATAFRAME"].columns),
                    "data": app.config["DATAFRAME"].to_dict(orient="records"),
                }
            )
        )

    @app.get("/api/catalog")
    def catalog() -> object:
        return jsonify(json_safe(app.config["CATALOG"]))

    @app.get("/api/jamii-tekelezi/locations")
    def jamii_tekelezi_locations() -> object:
        """Return JT counties, sub-counties, and full facility hierarchy for global filters."""
        jt_path = BASE_DIR / "data" / "jamii_tekelezi_filters.csv"
        if not jt_path.exists():
            return jsonify({"counties": [], "subcounties": [], "county_subcounties": {}, "facility_names": [], "facility_ids": [], "facility_id_name_map": {}, "facilities_by_subcounty": {}})
        try:
            jt_df = pd.read_csv(jt_path)
            counties = sorted(jt_df["county_name"].dropna().unique().tolist())
            subcounties = sorted(jt_df["subcounty_name"].dropna().unique().tolist())

            # Build facility ID ↔ name mapping
            id_name_records = jt_df[["facility_id", "facility_name"]].dropna().drop_duplicates()
            facility_id_name_map = dict(zip(id_name_records["facility_id"], id_name_records["facility_name"]))
            facility_names = sorted(id_name_records["facility_name"].unique().tolist())
            facility_ids = sorted(id_name_records["facility_id"].unique().tolist())

            # Build county → sub-counties mapping
            county_subcounties = {}
            for county in counties:
                subs = sorted(jt_df.loc[jt_df["county_name"] == county, "subcounty_name"].dropna().unique().tolist())
                county_subcounties[county] = subs

            # Build county → subcounty → [facility_id, facility_name] hierarchy
            facilities_by_subcounty = {}
            for _, row in jt_df.iterrows():
                c = row["county_name"]
                sc = row["subcounty_name"]
                fid = row["facility_id"]
                fname = row["facility_name"]
                if not (c and sc and fid):
                    continue
                key = f"{c}||{sc}"
                if key not in facilities_by_subcounty:
                    facilities_by_subcounty[key] = []
                facilities_by_subcounty[key].append({"id": fid, "name": fname})

            return jsonify({
                "counties": counties,
                "subcounties": subcounties,
                "county_subcounties": county_subcounties,
                "facility_names": facility_names,
                "facility_ids": facility_ids,
                "facility_id_name_map": facility_id_name_map,
                "facilities_by_subcounty": facilities_by_subcounty,
            })
        except Exception as e:
            return jsonify({"error": str(e), "counties": [], "subcounties": [], "county_subcounties": {}, "facility_names": [], "facility_ids": [], "facility_id_name_map": {}, "facilities_by_subcounty": {}})

    @app.get("/api/hiv-treatment/newly-started-art/by-county")
    def newly_started_art_by_county() -> object:
        """Return TX_NEW data for Jamii Tekelezi counties (Meru, Embu, Nyandarua, Tharaka Nithi).
        Reads the raw DHIS analytics CSV, maps wards to counties, filters to JT counties only.
        Also returns subcounty & facility filter data from the consolidated filters CSV.
        Supports ?month=YYYYMM filter.
        """
        try:
            # Load Jamii Tekelezi filter definitions
            jt_path = BASE_DIR / "data" / "jamii_tekelezi_filters.csv"
            if not jt_path.exists():
                return jsonify(json_safe({"error": "Jamii Tekelezi filters not found."})), 404
            jt_df = pd.read_csv(jt_path)
            jt_counties = sorted(jt_df["county_name"].unique())
            # Build subcounty list per county
            sc_by_county = {}
            for _, row in jt_df.iterrows():
                cn = row["county_name"]
                sn = row["subcounty_name"]
                if cn not in sc_by_county:
                    sc_by_county[cn] = set()
                sc_by_county[cn].add(sn)
            jt_subcounties = {c: sorted(v) for c, v in sc_by_county.items()}
            # Build facility list per subcounty
            fac_by_sc = {}
            for _, row in jt_df.iterrows():
                cn = row["county_name"]
                sn = row["subcounty_name"]
                fn = row["facility_name"]
                key = f"{cn}||{sn}"
                if key not in fac_by_sc:
                    fac_by_sc[key] = set()
                fac_by_sc[key].add(fn)
            jt_facilities = {k: sorted(v) for k, v in fac_by_sc.items()}

            # Load raw TX_NEW data
            raw_path = BASE_DIR / "data" / "dhis" / "raw" / "hiv_newly_started_art.csv"
            if not raw_path.exists():
                return jsonify(json_safe({"error": "Raw data not found."})), 404

            # Load org units for county mapping
            ou_path = BASE_DIR / "data" / "dhis" / "meta" / "organisation_units.csv"
            ou_df = pd.read_csv(ou_path)
            ou_name = dict(zip(ou_df["id"].astype(str), ou_df["name"]))
            ou_path_map = dict(zip(ou_df["id"].astype(str), ou_df["path"].astype(str)))

            # Load raw data and filter for TX_NEW
            df = pd.read_csv(raw_path)
            tx_new = df[df["dx"] == "gv7bbGesTTJ"].copy()

            # Map each ward to its county (level 2 in path)
            def get_county(oid: str) -> str:
                p = ou_path_map.get(oid, "")
                parts = p.split("/")
                if len(parts) >= 3:
                    return ou_name.get(parts[2], parts[2])
                return "Unknown"

            tx_new["county"] = tx_new["ou"].map(get_county)

            # Filter to JT counties only
            tx_new = tx_new[tx_new["county"].isin(jt_counties)]

            # Aggregate by county and month
            grouped = (
                tx_new.groupby(["county", "pe"], as_index=False)["value"]
                .sum()
                .rename(columns={"pe": "period", "value": "total"})
                .sort_values(["period", "county"])
            )

            # Format period label
            grouped["period_label"] = grouped["period"].apply(
                lambda p: f"{str(p)[:4]}-{str(p)[4:]}" if len(str(p)) == 6 else str(p)
            )

            # Build available months list (sorted newest first)
            months_list = sorted(tx_new["pe"].unique(), reverse=True)
            months = [
                {"period": str(m), "label": f"{str(m)[:4]}-{str(m)[4:]}"}
                for m in months_list
            ]

            # Optional month filter
            month_filter = request.args.get("month", "")
            if month_filter:
                grouped = grouped[grouped["period"] == month_filter]

            return jsonify(json_safe({
                "rows": grouped.to_dict(orient="records"),
                "months": months,
                "all_counties": jt_counties,
                "default_counties": jt_counties,  # all JT counties shown by default
                "default_month": str(months_list[0]) if months_list else "",
                "subcounties": jt_subcounties,
                "facilities": jt_facilities,
            }))
        except Exception as exc:
            return jsonify(json_safe({"error": str(exc)})), 500

    @app.get("/api/hiv-treatment/nart-trend")
    def nart_trend() -> object:
        """Return monthly trend for three NART metrics (Total, Males, Adults 15+)
        for a given Jamii Tekelezi county (default Meru County).
        """
        county_filter = (request.args.get("county") or "Meru County").strip()

        # ── Male Tx_New STA IDs (all age groups) ──
        MALE_DX = {
            "rpL7wMYNPDH", "s9iBEnfSHhh", "JprDjnAyB0f", "PG5Ynz9xGCu",
            "SNOcc1Tq2iH", "VqYNMLji5U5", "wJWCrZVh1iu", "jjXJNig8fxs",
            "Lbs5RUpnwPD", "cQrYHDWkY2y", "QRV2YRNGYJ6", "Mt9G8jCODUw",
            "FDRjPKGGVC9", "ShO7o3bHsNr", "ivLPgJtKgcN",
        }
        # ── Adults 15+ Tx_New STA IDs (both sexes) ──
        ADULT_15P_DX = MALE_DX - {"rpL7wMYNPDH", "s9iBEnfSHhh", "JprDjnAyB0f", "PG5Ynz9xGCu"} | {
            "NBMvd95wp7t", "X7QikQUsYB1", "MOqDhGiw7W6", "BFbmB3WxGPd",
            "xz2f0oONxQx", "f59E1kimKqe", "UjIzCVxESAz", "Y4jKOMblgII",
            "hUHq4KO9YMz", "tWTgIibsKJ5", "AFdikiNpC3e", "GMoRCzegC6C",
            "Bs5etPcLz7w", "Ps8a7Mv1xIn", "sQcd8UD8Mrs",
        }

        try:
            raw_path = BASE_DIR / "data" / "dhis" / "raw" / "hiv_newly_started_art.csv"
            ou_path = BASE_DIR / "data" / "dhis" / "meta" / "organisation_units.csv"
            if not raw_path.exists():
                return jsonify(json_safe({"error": "Raw data not found."})), 404

            df = pd.read_csv(raw_path)
            ou_df = pd.read_csv(ou_path)
            ou_name = dict(zip(ou_df["id"].astype(str), ou_df["name"]))
            ou_path_map = dict(zip(ou_df["id"].astype(str), ou_df["path"].astype(str)))

            def get_county(oid: str) -> str:
                p = ou_path_map.get(oid, "")
                parts = p.split("/")
                if len(parts) >= 3:
                    return ou_name.get(parts[2], parts[2])
                return "Unknown"

            df["county"] = df["ou"].map(get_county)
            df = df[df["county"] == county_filter]

            # Total TX_NEW
            total = df[df["dx"] == "gv7bbGesTTJ"]
            total_agg = total.groupby("pe", as_index=False)["value"].sum()

            # Males TX_NEW
            males = df[df["dx"].isin(MALE_DX)]
            males_agg = males.groupby("pe", as_index=False)["value"].sum()

            # Adults 15+
            adults = df[df["dx"].isin(ADULT_15P_DX)]
            adults_agg = adults.groupby("pe", as_index=False)["value"].sum()

            all_periods = sorted(set(
                list(total["pe"].unique()) +
                list(males["pe"].unique()) +
                list(adults["pe"].unique())
            ), reverse=True)

            def to_label(p):
                s = str(int(p))
                return f"{s[:4]}-{s[4:]}" if len(s) == 6 else s

            # Build trend series (chronological order)
            periods_chrono = sorted(all_periods)
            trend = []
            for p in periods_chrono:
                t_row = total_agg[total_agg["pe"] == p]
                m_row = males_agg[males_agg["pe"] == p]
                a_row = adults_agg[adults_agg["pe"] == p]
                trend.append({
                    "period": str(int(p)),
                    "label": to_label(p),
                    "total": round(float(t_row["value"].iloc[0]), 1) if len(t_row) else 0,
                    "males": round(float(m_row["value"].iloc[0]), 1) if len(m_row) else 0,
                    "adults_15plus": round(float(a_row["value"].iloc[0]), 1) if len(a_row) else 0,
                })

            return jsonify(json_safe({
                "county": county_filter,
                "metrics": [
                    {"key": "total", "label": "Newly Started on ART (Total)"},
                    {"key": "males", "label": "Males Started on ART"},
                    {"key": "adults_15plus", "label": "Adults Started on ART (15+ Yrs)"},
                ],
                "trend": trend,
                "periods": [{"period": str(int(p)), "label": to_label(p)} for p in periods_chrono],
            }))
        except Exception as exc:
            return jsonify(json_safe({"error": str(exc)})), 500

    @app.get("/api/facilities")
    def facilities() -> object:
        page = max(int(request.args.get("page", 1)), 1)
        page_size = min(max(int(request.args.get("page_size", 20)), 5), 100)
        search = (request.args.get("search") or "").strip().lower()
        payload = build_facility_page(
            app.config["DATAFRAME"],
            page=page,
            page_size=page_size,
            search=search,
            location_hierarchy=app.config.get("LOCATION_HIERARCHY"),
            hospitals_frame=app.config.get("HOSPITALS_TABLE"),
            org_units_map=app.config.get("ORG_UNITS_LOOKUP", {}),
            hospital_map=app.config.get("HOSPITALS_LOOKUP", {}),
        )
        return jsonify(json_safe(payload))

    @app.post("/api/chat")
    def chat() -> object:
        payload = request.get_json(silent=True) or {}
        question = (payload.get("question") or payload.get("message") or "").strip()
        chart_id = (payload.get("chart_id") or "").strip()

        if not question:
            return jsonify(json_safe({"error": "A question is required."})), 400

        cache_key = f"{chart_id}::{re.sub(r'\\s+', ' ', question.lower()).strip()}"
        with app.config["AI_RESPONSE_CACHE_LOCK"]:
            cached_response = app.config["AI_RESPONSE_CACHE"].get(cache_key)
        if cached_response is not None:
            return jsonify(json_safe(cached_response))

        try:
            sql_query, sql_source, ai_error = generate_sql(
                app.config["GEMINI_PROVIDERS"],
                app.config["GEMINI_ROUTER_STATE"],
                question,
                app.config["DATAFRAME"].columns.tolist(),
                app.config["SCHEMA_SQL"],
                chart_id=chart_id,
            )
            if sql_query is None:
                # No provider could produce a SQL query; surface the AI error to the client.
                return jsonify(json_safe({"error": f"AI providers failed: {ai_error}"})), 502
            validated_sql = validate_sql(sql_query, app.config["DATAFRAME"].columns.tolist())
            result_frame = run_safe_query(database, validated_sql)
            response_payload = build_chat_response(question, validated_sql, result_frame, sql_source, chart_id=chart_id)
            if ai_error:
                # include AI error details to help debugging in the UI
                response_payload["ai_error"] = str(ai_error)
            with app.config["AI_RESPONSE_CACHE_LOCK"]:
                app.config["AI_RESPONSE_CACHE"][cache_key] = response_payload
            return jsonify(json_safe(response_payload))
        except ValueError as exc:
            return jsonify(json_safe({"error": str(exc)})), 400
        except sqlite3.Error as exc:
            return jsonify(json_safe({"error": f"Database query failed: {exc}"})), 400
        except Exception as exc:  # pragma: no cover - defensive guard for production runtime
            return jsonify(json_safe({"error": f"Unexpected server error: {exc}"})), 500

    @app.get("/main.js")
    def main_js() -> object:
        return send_from_directory(BASE_DIR, "main.js")

    @app.post("/api/reload-csvs")
    def reload_csvs_endpoint() -> object:
        try:
            # call the scanner and return loaded files
            added = scan_and_load_additional_csvs()
            return jsonify(json_safe({"loaded_files": [str(p) for p in added]}))
        except Exception as exc:
            return jsonify(json_safe({"error": str(exc)})), 500

    @app.get("/api/ai-status")
    def ai_status() -> object:
        try:
            providers = app.config.get("GEMINI_PROVIDERS", [])
            router = app.config.get("GEMINI_ROUTER_STATE", {})
            cache_size = len(app.config.get("AI_RESPONSE_CACHE", {}))
            sdk_present = genai is not None
            return jsonify(
                json_safe(
                    {
                        "providers_count": len(providers),
                        "providers": providers,
                        "router_state": router,
                        "ai_response_cache_size": cache_size,
                        "sdk_present": bool(sdk_present),
                    }
                )
            )
        except Exception as exc:
            return jsonify(json_safe({"error": str(exc)})), 500

    @app.post("/api/clear-ai-cache")
    def clear_ai_cache() -> object:
        try:
            with app.config["AI_RESPONSE_CACHE_LOCK"]:
                app.config["AI_RESPONSE_CACHE"].clear()
            return jsonify(json_safe({"cleared": True, "ai_response_cache_size": 0}))
        except Exception as exc:
            return jsonify(json_safe({"error": str(exc)})), 500

    @app.get("/api/debug/check")
    def debug_check():
        return jsonify({"status": "alive", "has_superpower": HAS_SUPERPOWER})

    # ── Debug: list all registered routes ─────────────────────────────────
    @app.get("/api/debug/routes")
    def debug_routes():
        rules = sorted([r.rule for r in app.url_map.iter_rules()])
        return jsonify({"routes": rules, "count": len(rules)})

    # ── Helper: fetch from DHIS2 (direct HTTP) ─────────────────────────
    def _dhis2_fetch(dx_ids, ou_id, pe="LAST_12_MONTHS", coc_ids=None):
        """Fetch analytics rows from DHIS2.
        Returns:
          - Without coc_ids: dict of {period_label: summed_value}
          - With coc_ids: dict of {(period_label, coc_id): value}
        dx_ids can be a single string or a list/set of strings.
        """
        import requests as _req
        from requests.auth import HTTPBasicAuth

        # Always use the CHAK DHIS2 server
        dhis_base = "http://ereporting.chak.or.ke:8500/api"
        url_base = dhis_base + "/analytics.json"
        username = os.getenv("DHIS_USERNAME", "Johnbrian")
        password = os.getenv("DHIS_PASSWORD", "JOHNb123\\")
        auth = HTTPBasicAuth(username, password)

        if isinstance(dx_ids, (list, set, tuple)):
            dx_str = ";".join(dx_ids)
        else:
            dx_str = dx_ids

        if not dx_str or not dx_str.strip():
            return {}  # no DX IDs to query

        # Build dimensions – ou_id can be a single ID or list
        if isinstance(ou_id, (list, set, tuple)):
            ou_str = ";".join(ou_id)
        else:
            ou_str = ou_id
        dimensions = [f"dx:{dx_str}", f"pe:{pe}", f"ou:{ou_str}"]
        if coc_ids:
            if isinstance(coc_ids, (list, set, tuple)):
                coc_str = ";".join(coc_ids)
            else:
                coc_str = coc_ids
            dimensions.append(f"co:{coc_str}")

        # Try superpower first if available
        if HAS_SUPERPOWER:
            api_url = f"{url_base}?" + "&".join(
                f"dimension={d}" for d in dimensions
            ) + "&displayProperty=NAME"
            try:
                result = _superpower_fetch_result(api_url)
                if result.get("ok") and result.get("rows"):
                    return _parse_dhis2_rows(result["rows"], result.get("metaData", {}).get("items", {}), coc_ids)
            except Exception:
                pass  # fall through to direct HTTP

        # Direct HTTP fallback
        params = {
            "dimension": dimensions,
            "displayProperty": "NAME",
        }
        resp = _req.get(url_base, params=params, auth=auth, timeout=120)
        if not resp.ok:
            return {}
        data = resp.json()
        rows = data.get("rows", [])
        meta = data.get("metaData", {}).get("items", {})
        return _parse_dhis2_rows(rows, meta, coc_ids)

    def _parse_dhis2_rows(rows, meta, coc_ids=None):
        """Parse DHIS2 analytics rows into a dict."""
        if not rows:
            return {}
        out = {}
        if coc_ids:
            # COC-disaggregated: keyed by (period, coc_id)
            # Row format: [dx, coc, pe, ou, value] when ou dimension is present
            for row in rows:
                if len(row) < 4:
                    continue
                coc = str(row[1]) if len(row) > 1 else ""
                pe_code = str(row[2]) if len(row) > 2 else ""
                pe_name = meta.get(pe_code, {}).get("name", pe_code)
                val = float(row[-1]) if row[-1] else 0
                out[(pe_name, coc)] = out.get((pe_name, coc), 0) + val
        else:
            # Simple period-summed
            for row in rows:
                pe_code = str(row[1]) if len(row) > 1 else ""
                pe_name = meta.get(pe_code, {}).get("name", pe_code)
                val = float(row[-1]) if row[-1] else 0
                out[pe_name] = out.get(pe_name, 0) + val
        return out

    # ── Helper: find DX IDs from dictionary by pattern ─────────────────
    def _find_dx_by_pattern(prefix_pattern, age_bands):
        """Find male & female DX IDs from the data element dictionary.
        prefix_pattern: e.g. 'Tx_New STA' or 'TX_Curr STA'
        age_bands: list of strings like ['<1','1-4',...]
        Returns (males_list, females_list) of 11-char UIDs each.
        """
        try:
            df_elements, _ = _superpower_load_dict()
        except Exception:
            return [], []
        names = df_elements["name"].astype(str)
        males, females = [], []
        for age in age_bands:
            # Handle both "Tx_New STA <1M" (no comma) and "TX_Curr STA <1,M" (comma)
            pat_m = rf"{prefix_pattern}\s+{age},?\s*M\b"
            pat_f = rf"{prefix_pattern}\s+{age},?\s*F\b"
            m_match = df_elements[names.str.match(pat_m, case=False, na=False)]
            f_match = df_elements[names.str.match(pat_f, case=False, na=False)]
            if not m_match.empty:
                males.append(m_match.iloc[0]["id"])
            if not f_match.empty:
                females.append(f_match.iloc[0]["id"])
        return males, females

    # ── Indicator specs for each subtab type ───────────────────────────
    # DX IDs verified against dictionaries/master_data_elements.csv (2025-06-28)
    # ── COC IDs for "Finer Age Bands and Gender" category combo ──────
    # Age bands → Category Option Combo IDs (verified against CHAK DHIS2)
    _COC_MALES = [
        "AwerOu6rx5q","g2zP3yNwOOa","WTfu1bBSG12","X65JamO5tyb","hKHprPKwjL6",
        "uSDHHGh2DZo","sLaLEIDVusT","b91xfEPrY4D","EU7hVFz5Yyt","AR2E4Yiuo8Z",
        "dfkyp7ZQZSr","lswMoqT008e","Z6zV5L8i14I","XIc55yRW4aQ","g5bVF4b8hmV",
    ]  # <1,M  1-4,M  5-9,M  10-14,M  15-19,M  20-24,M  25-29,M  30-34,M  35-39,M  40-44,M  45-49,M  50-54,M  55-59,M  60-64,M  65+,M
    _COC_FEMALES = [
        "dcv8Lowu94w","D2aMSzo7SEw","HIS0TcFAoo8","Rr3uh3eAvKi","DYDpnZWu1XK",
        "m7Y0ddB212k","qy1vJGvFJeB","sk5UiD3PrxH","Vb7KzTvF83C","dchngmvBGvb",
        "VP1zCgdzuBb","uefSjW3VtZr","llt7APqVWyq","gs3y2muDLIK","YAtW6LDL24J",
    ]  # <1,F  1-4,F  5-9,F  10-14,F  15-19,F  20-24,F  25-29,F  30-34,F  35-39,F  40-44,F  45-49,F  50-54,F  55-59,F  60-64,F  65+,F

    INDICATOR_SPECS = {
        "tx_new": {
            "title": "Newly Started on ART",
            "aggregate": "vTTEybkXZ53",  # TX_NEW: Starting ART (CHAK DHIS2)
            "male_cocs": _COC_MALES,
            "female_cocs": _COC_FEMALES,
            "age_bands": ["<1","1-4","5-9","10-14","15-19","20-24",
                          "25-29","30-34","35-39","40-44","45-49",
                          "50-54","55-59","60-64","65+"],
            "color_total": "#2563eb",
            "color_male": "#10b981",
            "color_female": "#ec4899",
        },
        "tx_curr": {
            "title": "Current on ART",
            "aggregate": "kgzd9LfXZXq",  # TX_CURR (CHAK DHIS2)
            "male_cocs": _COC_MALES,
            "female_cocs": _COC_FEMALES,
            "age_bands": ["<1","1-4","5-9","10-14","15-19","20-24",
                          "25-29","30-34","35-39","40-44","45-49",
                          "50-54","55-59","60-64","65+"],
            "color_total": "#7c3aed",
            "color_male": "#10b981",
            "color_female": "#ec4899",
        },
        "vl": {
            "title": "VL Monitoring",
            "aggregate": "JGd3MwmKBuM",  # TX_PVLS (D) Routine
            "color_total": "#0891b2",
            "vl_pvls_d": "JGd3MwmKBuM",
            "vl_pvls_n": "FloZph8hN9z",   # TX_PVLS (N) Routine
        },
    }

    # ── JTP (Jamii Tekelezi Program) Treatment subtab specs ────────────
    # Data elements from JTP Monthly HIV Care and Treatment dataset
    JTP_SPECS = {
        "art_optimization": {
            "title": "ART Optimization",
            "metrics": {
                "regimen_1st_line": {"ids": ["zZGNba5d34c"], "label": "TX_CURR on 1st Line"},
                "regimen_2nd_line": {"ids": ["F0xtjHxDZ2e"], "label": "TX_CURR on 2nd Line"},
                "regimen_3rd_line": {"ids": ["Pk1PMmG4ml7"], "label": "TX_CURR on 3rd Line"},
                "on_dtg": {"ids": ["s62uidROGjG"], "label": "TX_CURR on DTG"},
                "eligible_dtg": {"ids": ["bsQdHW8sJ4b"], "label": "Eligible for DTG"},
                "efv_600": {"ids": ["lr1YorhNrJT"], "label": "Active on EFV-600"},
                "efv_400": {"ids": ["ggO3YzjB9j4"], "label": "Active on EFV-400"},
                "pi_based": {"ids": ["Z4g3jskQn9c"], "label": "Active on PI Regimen"},
                "viremia_clinic": {"ids": ["JGIZOGP6bGU"], "label": "Active in Viremia Clinic"},
            },
        },
        "dsd": {
            "title": "Differentiated Service Delivery (DSD)",
            "metrics": {
                "eligible_dc": {"ids": ["zkZIEcm0mFs"], "label": "Eligible for DC"},
                "on_dc": {"ids": ["qfLHg1lbN3W"], "label": "Clients Put on DC"},
                "eligible_dcm": {"ids": ["oJfkiD0C599"], "label": "Eligible for DCM"},
                "dcm_community": {"ids": ["oHTbEMaKn7L"], "label": "DCM - Community"},
                "dcm_facility": {"ids": ["cPEthRR8Zs1"], "label": "DCM - Facility"},
                "arv_dispensing": {"ids": ["Lo3GoG3lxOF"], "label": "ARV Dispensing"},
            },
        },
        "treatment_outcomes": {
            "title": "Treatment Outcomes",
            "metrics": {
                "iit_3m": {"ids": ["Cn2q8OMIHDD"], "label": "IIT > 3 Months"},
                "iit_1m": {"ids": ["Gg3ZzBADtz8"], "label": "IIT <= 3 Months"},
                "stopped": {"ids": ["HjRScpxwVQn"], "label": "Stopped Treatment"},
                "died": {"ids": ["DrtGJ1cgA3J"], "label": "Died this Month"},
                "transfers_out": {"ids": ["YX65zEg6s5R"], "label": "Transferred Out"},
                "transfers_in": {"ids": ["vODolOs8eBi"], "label": "Transferred In"},
                "rtt": {"ids": ["tCFth5mfGz5"], "label": "Return to Treatment (RTT)"},
            },
        },
        "otz": {
            "title": "OTZ (O and Teen Club)",
            "metrics": {
                "booked": {"ids": ["ep6NJHQ9LJa"], "label": "OTZ Booked"},
                "adher_95": {"ids": ["kMe6SRtZ6pn"], "label": "Adherence >95%"},
                "basevl_lt1000": {"ids": ["O6wVcguUXbQ"], "label": "Baseline VL <1000"},
                "basevl_lt200": {"ids": ["caFpNcJu2q2"], "label": "Baseline VL <200"},
                "basevl_ldl": {"ids": ["JFfFnMNV2dL"], "label": "Baseline VL LDL"},
            },
        },
        "ovc": {
            "title": "OVC (Orphans and Vulnerable Children)",
            "metrics": {
                "calhiv_enrolled": {"ids": ["Ml7NYiKXKJ9"], "label": "CALHIV OVC Enrolled"},
                "prev_15_24": {"ids": ["d7PKefTNRPT"], "label": "ICT PREV 15-24yrs"},
                "prev_lt15": {"ids": ["jgtv4fG2Vpm"], "label": "ICT PREV <15yrs"},
                "prev_gt25": {"ids": ["s8sJolX8W8I"], "label": "ICT PREV >25yrs"},
                "ovc_hivstat": {"ids": ["vd42QeUMvb2"], "label": "OVC HIV Status Known"},
                "ovc_serv": {"ids": ["CNdhjjVHiHp"], "label": "OVC Services"},
            },
        },
        "covid": {
            "title": "COVID-19",
            "metrics": {
                "screened": {"ids": ["z4ylGRyQ7Rl"], "label": "CCC Screened for COVID"},
            },
        },
        "ahd": {
            "title": "Advanced HIV Disease (AHD)",
            "metrics": {
                "kepi": {"ids": ["g0s65Dm47CA"], "label": "CALHIV KEPI"},
                "not_imm": {"ids": ["oWSf2zNhstC"], "label": "CALHIV Not Immunized"},
                "pneumo": {"ids": ["vFByuxLPFX7"], "label": "CALHIV Pneumo"},
                "rota": {"ids": ["iBsCVgCGyQ7"], "label": "CALHIV Rota"},
                "cd4_smearpos": {"ids": ["R5KyL0fYJle"], "label": "CD4 TB Smear+"},
                "iit_cd4": {"ids": ["KRvAhAQ5O46"], "label": "IIT CD4"},
                "iit_cd4_lt200": {"ids": ["rrfFv94aqgy"], "label": "IIT CD4 <200"},
            },
        },
        "adverse_events": {
            "title": "Adverse Events (AE)",
            "metrics": {
                "cod_tb": {"ids": ["LIvYFVNx00M"], "label": "COD - TB"},
                "cod_cancer": {"ids": ["RlKpuE0qv0q"], "label": "COD - Cancer"},
                "cod_other_hiv": {"ids": ["ZnmeKVBraIA"], "label": "COD - Other HIV Disease"},
                "cod_other_natural": {"ids": ["hgV3WXdiPbZ"], "label": "COD - Other Natural Causes"},
                "cod_non_natural": {"ids": ["Lw7SrH10igs"], "label": "COD - Non Natural Causes"},
                "cod_unknown": {"ids": ["yrERSv923Qh"], "label": "COD - Unknown Causes"},
            },
        },
    }

    # ── HTS (HIV Testing Services) indicator specs ────────────────────
    HTS_SPECS = {
        "hts_uptake": {
            "title": "HIV Testing Services Uptake",
            "metrics": {
                "hts_tested": {
                    "ids": ["ymKviaHZtQN","vFlUDposW0Y","XKAlilawdhN","THJbtDzxplR",
                            "Lwtqyjus0Mb","QBsyLQZRdiH","XYhYAMivUX5","J4zibSjbBCt"],
                    "label": "HTS TST Numerator",
                },
                "hts_positive": {
                    "ids": ["CcOr3MB7Mh4"],  # MOH731_HV01-19: Total HIV Positive
                    "label": "HIV Positive",
                },
            },
            # positivity_rate computed in endpoint
        },
        "hts_linkage": {
            "title": "HIV Testing Services Linkage",
            "metrics": {
                "linked_within": {
                    "ids": ["wQ5AA7GTs9G","YroUdlNVeR2","h13L1gcUaCS"],
                    "label": "Linked Within Facility",
                },
                "linked_outside": {
                    "ids": ["DdPzCAtN3J2","ZnetI7sd8Ub","BeO9dmxTBMg"],
                    "label": "Linked Outside Facility",
                },
                "total_tested": {
                    "ids": ["ymKviaHZtQN","vFlUDposW0Y","XKAlilawdhN","THJbtDzxplR",
                            "Lwtqyjus0Mb","QBsyLQZRdiH","XYhYAMivUX5","J4zibSjbBCt"],
                    "label": "Total Tested (Ref)",
                },
            },
        },
        "partner_notification": {
            "title": "Partner Notification Services",
            "metrics": {
                "index_offered": {
                    "ids": ["gp0RYhjsc1f"],
                    "label": "Index Clients Offered PNS",
                },
                "index_accepted": {
                    "ids": ["sTBggmuuiGR"],
                    "label": "Index Clients Accepting PNS",
                },
                "contacts_tested": {
                    "ids": ["DtJ8Kpaquhx"],
                    "label": "Contacts Tested",
                },
            },
        },
        "prep": {
            "title": "PrEP",
            "metrics": {
                "prep_new": {
                    "ids": ["HmUEZ2yWtAE","tSOqRYW3fUp","CYLF8hUOHpv","Q57YuHsnTKm",
                            "OOhFACMqmKp","hxfjIrnxHBF","EmzN6C78vFE","BSx4nKKwK1r",
                            "mbSrJM6OvQo","N3IsvP0sUF5","DQR7sycvi6V","JJsuQUWLYsD",
                            "qvhr1STgYAD","N6iP1PPLmyX","EX7lZNXZXDe"],
                    "label": "PrEP New",
                },
                "prep_curr": {
                    "ids": ["UhOTGSCLcvz","ENYVWNfmlWi","m0sXEku2oeB","McHzszUZFtf"],
                    "label": "PrEP Current",
                },
                "prep_screened": {
                    "ids": ["Th1loMyxBhR","C0z7kQZw2LZ","dc0q1SbjyPU"],
                    "label": "PrEP Screened",
                },
            },
        },
    }

    # ── Unified DHIS2 Live Query for HIV Treatment subtabs ────────────
    @app.get("/api/hiv-treatment/dhis-live")
    def hiv_treatment_dhis_live() -> object:
        """Unified DHIS2 endpoint for all HIV Treatment subtabs.
        Params: ?type=tx_new|tx_curr|vl&county=...&subcounty=...&facility=...&period=LAST_12_MONTHS
        """

        qtype = (request.args.get("type") or "tx_new").strip().lower()
        county = (request.args.get("county") or "Meru County").strip()
        pe = (request.args.get("period") or "LAST_12_MONTHS").strip()
        sc_filter = request.args.get("subcounty", "").strip()
        fac_filter = request.args.get("facility", "").strip()

        # ── Check both INDICATOR_SPECS and JTP_SPECS ──
        spec = INDICATOR_SPECS.get(qtype)
        jtp_spec = JTP_SPECS.get(qtype) if not spec else None
        if not spec and not jtp_spec:
            return jsonify(json_safe({
                "error": f"Unknown type '{qtype}'. Use: tx_new, tx_curr, vl, art_optimization, dsd, treatment_outcomes, otz, ovc, covid, ahd, adverse_events"
            })), 400

        ou_id, is_multi_ou = _resolve_ou_ids(county, sc_filter or None, fac_filter or None)

        # ── JTP multi-metric type handling ──
        if jtp_spec:
            title = jtp_spec["title"]
            from concurrent.futures import ThreadPoolExecutor
            errors = []
            metrics_data = {}
            with ThreadPoolExecutor(max_workers=8) as ex:
                future_map = {}
                for mkey, mmeta in jtp_spec["metrics"].items():
                    dx_str = ";".join(mmeta["ids"])
                    fut = ex.submit(_dhis2_fetch, dx_str, ou_id, pe, None)
                    future_map[fut] = mkey
                for fut in future_map:
                    mkey = future_map[fut]
                    try:
                        metrics_data[mkey] = fut.result()
                    except Exception as exc:
                        errors.append(f"{jtp_spec['metrics'][mkey]['label']}: {exc}")
                        metrics_data[mkey] = {}

            all_set = set()
            for md in metrics_data.values():
                all_set.update(md.keys())
            all_periods = sorted(all_set, key=_period_sort_key)

            def _jtp_label(p):
                return f"{p[:4]}-{p[4:]}" if len(str(p)) == 6 else str(p)

            trend = []
            for p in all_periods:
                entry = {"period": p, "label": _jtp_label(p)}
                for mkey in jtp_spec["metrics"]:
                    entry[mkey] = round(float(metrics_data.get(mkey, {}).get(p, 0)), 1)
                trend.append(entry)

            metric_list = []
            for mkey, mmeta in jtp_spec["metrics"].items():
                metric_list.append({"key": mkey, "label": mmeta["label"]})

            return jsonify(json_safe({
                "type": qtype,
                "title": title,
                "source": "dhis2_live",
                "county": county,
                "subcounty": sc_filter or None,
                "facility": fac_filter or None,
                "ou_id": ou_id,
                "period_range": pe,
                "metrics": metric_list,
                "trend": trend,
                "monthly_cards": [],
                "fetched_at": time.strftime("%Y-%m-%d %H:%M:%S"),
                "errors": errors if errors else None,
            }))

        title = spec["title"]

        # ── Get male/female COC IDs from spec ──
        male_cocs = spec.get("male_cocs", [])
        female_cocs = spec.get("female_cocs", [])
        age_bands = spec.get("age_bands", [])

        # ── Prepare parallel fetch: aggregate + COC-disaggregated ──
        from concurrent.futures import ThreadPoolExecutor, as_completed

        fetch_specs = []
        # 1) Aggregate (single DX, no COC)
        fetch_specs.append(("total", spec["aggregate"]))

        if qtype in ("tx_new", "tx_curr"):
            # 2) Single COC-disaggregated query for all 30 age×sex combos
            all_cocs = male_cocs + female_cocs
            fetch_specs.append(("coc_data", spec["aggregate"]))
        elif qtype == "vl":
            # Fetch TX_PVLS (D) — the numerator
            pvls_d_id = spec.get("vl_pvls_d", "")
            if pvls_d_id:
                fetch_specs.append(("pvls_d_raw", pvls_d_id))
            # Fetch TX_CURR — the denominator
            tx_curr_spec = INDICATOR_SPECS.get("tx_curr", {})
            tx_curr_agg = tx_curr_spec.get("aggregate", "")
            if tx_curr_agg:
                fetch_specs.append(("tx_curr", tx_curr_agg))

        # ── Parallel fetch ──
        fetched = {}
        errors = []
        with ThreadPoolExecutor(max_workers=6) as executor:
            futures = {}
            for key, dx in fetch_specs:
                cocs_for_key = None
                if key == "coc_data":
                    cocs_for_key = all_cocs
                futures[executor.submit(_dhis2_fetch, dx, ou_id, pe, cocs_for_key)] = key
            for future in as_completed(futures):
                key = futures[future]
                try:
                    fetched[key] = future.result()
                except Exception as exc:
                    errors.append(f"{key}: {exc}")
                    fetched[key] = {}

        # ── Build trend data ──
        total_data = fetched.get("total", {})
        all_periods = sorted(set(total_data.keys()), key=_period_sort_key)

        trend = []
        monthly_cards = []

        if qtype in ("tx_new", "tx_curr"):
            # COC-disaggregated data: dict of (period, coc_id) → value
            coc_data = fetched.get("coc_data", {})
            male_coc_set = set(male_cocs)
            female_coc_set = set(female_cocs)
            coc_to_age_idx = {}
            for i, cid in enumerate(male_cocs):
                coc_to_age_idx[cid] = ("male", i)
            for i, cid in enumerate(female_cocs):
                coc_to_age_idx[cid] = ("female", i)

            def to_label(p):
                return f"{p[:4]}-{p[4:]}" if len(str(p)) == 6 else str(p)

            for p in all_periods:
                total_val = round(float(total_data.get(p, 0)), 1)

                # Sum males and females from COC rows for this period
                male_val = 0.0
                female_val = 0.0
                male_band_sums = [0.0] * len(age_bands)
                female_band_sums = [0.0] * len(age_bands)

                for (cp, cid), val in coc_data.items():
                    if cp != p:
                        continue
                    info = coc_to_age_idx.get(cid)
                    if not info:
                        continue
                    sex, idx = info
                    if sex == "male":
                        male_val += val
                        if idx < len(male_band_sums):
                            male_band_sums[idx] += val
                    else:
                        female_val += val
                        if idx < len(female_band_sums):
                            female_band_sums[idx] += val

                male_bands = [{"age": a, "value": round(v, 1)} for a, v in zip(age_bands, male_band_sums)]
                female_bands = [{"age": a, "value": round(v, 1)} for a, v in zip(age_bands, female_band_sums)]

                trend.append({
                    "period": p,
                    "label": to_label(p),
                    "total": total_val,
                    "males": male_val,
                    "females": female_val,
                })
                monthly_cards.append({
                    "period": p,
                    "label": to_label(p),
                    "total": total_val,
                    "males": male_val,
                    "females": female_val,
                    "male_bands": male_bands,
                    "female_bands": female_bands,
                })

            return jsonify(json_safe({
                "type": qtype,
                "title": title,
                "source": "dhis2_live",
                "county": county,
                "subcounty": sc_filter or None,
                "facility": fac_filter or None,
                "ou_id": ou_id,
                "period_range": pe,
                "metrics": [
                    {"key": "total", "label": f"{title} (Total)",
                     "color": spec["color_total"]},
                    {"key": "males", "label": f"Males",
                     "color": spec["color_male"]},
                    {"key": "females", "label": f"Females",
                     "color": spec["color_female"]},
                ],
                "age_bands": age_bands,
                "trend": trend,
                "monthly_cards": monthly_cards,
                "fetched_at": time.strftime("%Y-%m-%d %H:%M:%S"),
                "errors": errors if errors else None,
            }))

        elif qtype == "vl":
            def to_label(p):
                return f"{p[:4]}-{p[4:]}" if len(str(p)) == 6 else str(p)

            pvls_d_data = fetched.get("pvls_d_raw", {})
            tx_curr_data = fetched.get("tx_curr", {})
            for p in all_periods:
                pvls_d_val = float(pvls_d_data.get(p, 0)) or 0
                tx_curr_val = float(tx_curr_data.get(p, 0)) or 0
                vl_uptake = round(
                    (pvls_d_val / tx_curr_val * 100) if tx_curr_val > 0 else 0, 1
                )
                entry = {
                    "period": p,
                    "label": to_label(p),
                    "pvls_d": pvls_d_val,
                    "tx_curr": tx_curr_val,
                    "vl_uptake": vl_uptake,
                }
                trend.append(entry)
                monthly_cards.append(entry)

            metric_list = [
                {"key": "pvls_d", "label": "TX_PVLS (D)"},
                {"key": "tx_curr", "label": "TX_CURR"},
                {"key": "vl_uptake", "label": "% VL Uptake"},
            ]

            return jsonify(json_safe({
                "type": qtype,
                "title": title,
                "source": "dhis2_live",
                "county": county,
                "subcounty": sc_filter or None,
                "facility": fac_filter or None,
                "ou_id": ou_id,
                "period_range": pe,
                "metrics": metric_list,
                "trend": trend,
                "monthly_cards": monthly_cards,
                "fetched_at": time.strftime("%Y-%m-%d %H:%M:%S"),
                "errors": errors if errors else None,
            }))

    # ── Unified DHIS2 Live Query for HIV Testing subtabs ──────────────
    @app.get("/api/hiv-testing/dhis-live")
    def hiv_testing_dhis_live() -> object:
        """Unified DHIS2 endpoint for all HIV Testing subtabs.
        Params: ?type=hts_uptake|hts_linkage|partner_notification|prep
                &county=...&subcounty=...&facility=...
        Delivers 3-metric multi-line chart data per subtab.
        """
        qtype = (request.args.get("type") or "hts_uptake").strip().lower()
        county = (request.args.get("county") or "Meru County").strip()
        pe = (request.args.get("period") or "LAST_12_MONTHS").strip()
        sc_filter = request.args.get("subcounty", "").strip()
        fac_filter = request.args.get("facility", "").strip()

        spec = HTS_SPECS.get(qtype)
        if not spec:
            return jsonify(json_safe({
                "error": f"Unknown type '{qtype}'. Use: hts_uptake, hts_linkage, partner_notification, prep"
            })), 400

        title = spec["title"]
        ou_id, _ = _resolve_ou_ids(county, sc_filter or None, fac_filter or None)

        from concurrent.futures import ThreadPoolExecutor

        errors = []
        metrics_data = {}

        # Fetch each metric in parallel
        with ThreadPoolExecutor(max_workers=6) as ex:
            future_map = {}
            for mkey, mmeta in spec["metrics"].items():
                dx_str = ";".join(mmeta["ids"])
                fut = ex.submit(_dhis2_fetch, dx_str, ou_id, pe, None)
                future_map[fut] = mkey
            for fut in future_map:
                mkey = future_map[fut]
                try:
                    metrics_data[mkey] = fut.result()
                except Exception as exc:
                    errors.append(f"{spec['metrics'][mkey]['label']}: {exc}")
                    metrics_data[mkey] = {}

        # Collect all period keys across all metrics
        all_set = set()
        for md in metrics_data.values():
            all_set.update(md.keys())
        all_periods = sorted(all_set, key=_period_sort_key)

        def _hts_label(p):
            return f"{p[:4]}-{p[4:]}" if len(str(p)) == 6 else str(p)

        trend = []
        for p in all_periods:
            entry = {"period": p, "label": _hts_label(p)}
            for mkey in spec["metrics"]:
                entry[mkey] = round(
                    float(metrics_data.get(mkey, {}).get(p, 0)), 1
                )
            # Compute positivity rate for hts_uptake
            if qtype == "hts_uptake":
                tested = float(metrics_data.get("hts_tested", {}).get(p, 0))
                positive = float(metrics_data.get("hts_positive", {}).get(p, 0))
                entry["positivity_rate"] = round(
                    (positive / tested * 100) if tested > 0 else 0, 1
                )
            trend.append(entry)

        metric_list = [
            {"key": mk, "label": mm["label"]}
            for mk, mm in spec["metrics"].items()
        ]
        if qtype == "hts_uptake":
            metric_list.append({
                "key": "positivity_rate",
                "label": "HTS TST % Positive",
                "is_pct": True,
            })

        return jsonify(json_safe({
            "type": qtype,
            "title": title,
            "source": "dhis2_live",
            "county": county,
            "subcounty": sc_filter or None,
            "facility": fac_filter or None,
            "ou_id": ou_id,
            "period_range": pe,
            "metrics": metric_list,
            "trend": trend,
            "fetched_at": time.strftime("%Y-%m-%d %H:%M:%S"),
            "errors": errors if errors else None,
        }))

    # ── DHIS2 Live Query: NART Trend via Superpower ────────────────────
    @app.get("/api/hiv-treatment/nart-dhis-live")
    def nart_dhis_live() -> object:
        """Query DHIS2 via the superpower module for 3 NART metrics.
        Routes the natural-language question through ai_translator's
        generate_dhis2_url + fetch_query_result pipeline.
        Params: ?county=...&subcounty=...&facility=...
        """
        county = (request.args.get("county") or "Meru County").strip()
        sc_filter = request.args.get("subcounty", "").strip()
        fac_filter = request.args.get("facility", "").strip()
        ou_id, is_multi_ou = _resolve_ou_ids(county, sc_filter or None, fac_filter or None)

        # ── 3 metric groups ──
        DX_TOTAL = "gv7bbGesTTJ"
        MALE_DX = [
            "rpL7wMYNPDH","s9iBEnfSHhh","JprDjnAyB0f","PG5Ynz9xGCu",
            "SNOcc1Tq2iH","VqYNMLji5U5","wJWCrZVh1iu","jjXJNig8fxs",
            "Lbs5RUpnwPD","cQrYHDWkY2y","QRV2YRNGYJ6","Mt9G8jCODUw",
            "FDRjPKGGVC9","ShO7o3bHsNr","ivLPgJtKgcN",
        ]
        ADULT_15P_DX = (
            set(MALE_DX) - {"rpL7wMYNPDH","s9iBEnfSHhh","JprDjnAyB0f","PG5Ynz9xGCu"}
        ) | {
            "NBMvd95wp7t","X7QikQUsYB1","MOqDhGiw7W6","BFbmB3WxGPd",
            "xz2f0oONxQx","f59E1kimKqe","UjIzCVxESAz","Y4jKOMblgII",
            "hUHq4KO9YMz","tWTgIibsKJ5","AFdikiNpC3e","GMoRCzegC6C",
            "Bs5etPcLz7w","Ps8a7Mv1xIn","sQcd8UD8Mrs",
        }

        dhis_base = os.getenv("DHIS_BASE_URL") or "http://ereporting.chak.or.ke:8500/api/"
        url_base = dhis_base.rstrip("/") + "/analytics.json"
        pe = "LAST_12_MONTHS"

        # ── Step 1: Superpower generates URL for the TOTAL metric ──
        superpower_url = None
        if HAS_SUPERPOWER:
            try:
                question = (
                    f"Tx_New STA for {county} over the last 12 months"
                )
                superpower_url = _superpower_generate_url(question)
                if superpower_url:
                    print(f"[Superpower] Generated URL: {superpower_url[:120]}...")
            except Exception as exc:
                print(f"[Superpower] generate_dhis2_url failed: {exc}")

        # ── Step 2: Superpower fetch for all 3 metric groups ──
        def fetch_via_superpower(dx_ids):
            """Build URL & fetch through superpower's fetch_query_result."""
            if isinstance(dx_ids, (list, set, tuple)):
                dx_str = ";".join(dx_ids)
            else:
                dx_str = dx_ids
            if isinstance(ou_id, (list, set, tuple)):
                ou_str = ";".join(ou_id)
            else:
                ou_str = ou_id
            api_url = (
                f"{url_base}?"
                f"dimension=dx:{dx_str}"
                f"&dimension=pe:{pe}"
                f"&dimension=ou:{ou_str}"
                f"&displayProperty=NAME"
            )
            if HAS_SUPERPOWER:
                result = _superpower_fetch_result(api_url)
                if result.get("ok") and result.get("rows"):
                    return result["rows"]
                # fallback: superpower failed, use direct
            import requests as _req
            from requests.auth import HTTPBasicAuth
            username = os.getenv("DHIS_USERNAME", "Johnbrian")
            password = os.getenv("DHIS_PASSWORD", "JOHNb123\\")
            auth = HTTPBasicAuth(username, password)
            resp = _req.get(api_url, auth=auth, timeout=120)
            if not resp.ok:
                return []
            data = resp.json()
            hdrs = [h.get("name","").lower() for h in data.get("headers",[])]
            pe_i = next((i for i,h in enumerate(hdrs) if h in ("pe","period")), 0)
            val_i = next((i for i,h in enumerate(hdrs) if h=="value"), len(hdrs)-1)
            rows = data.get("rows", [])
            meta = data.get("metaData",{}).get("items",{})
            out = []
            for row in rows:
                pe_name = meta.get(str(row[pe_i]),{}).get("name", str(row[pe_i]))
                out.append({"period": pe_name, "value": float(row[val_i]) if row[val_i] else 0})
            return out

        def aggregate_by_period(rows):
            """Sum values by period from superpower-style rows."""
            result = {}
            for r in rows:
                p = str(r.get("period", ""))
                v = float(r.get("value", 0))
                result[p] = result.get(p, 0) + v
            return result

        try:
            total_rows = fetch_via_superpower(DX_TOTAL)
            males_rows = fetch_via_superpower(MALE_DX)
            adults_rows = fetch_via_superpower(ADULT_15P_DX)

            total_data = aggregate_by_period(total_rows)
            males_data = aggregate_by_period(males_rows)
            adults_data = aggregate_by_period(adults_rows)

            all_periods = sorted(set(
                list(total_data.keys()) +
                list(males_data.keys()) +
                list(adults_data.keys())
            ), key=_period_sort_key)

            def to_label(p):
                return f"{p[:4]}-{p[4:]}" if len(str(p)) == 6 else str(p)

            trend = []
            for p in all_periods:
                trend.append({
                    "period": p,
                    "label": to_label(p),
                    "total": round(float(total_data.get(p, 0)), 1),
                    "males": round(float(males_data.get(p, 0)), 1),
                    "adults_15plus": round(float(adults_data.get(p, 0)), 1),
                })

            return jsonify(json_safe({
                "source": "dhis2_superpower",
                "county": county,
                "subcounty": sc_filter or None,
                "facility": fac_filter or None,
                "ou_id": ou_id,
                "period_range": pe,
                "superpower_url": superpower_url,
                "metrics": [
                    {"key": "total", "label": "Newly Started on ART (Total)"},
                    {"key": "males", "label": "Males Started on ART"},
                    {"key": "adults_15plus", "label": "Adults Started on ART (15+ Yrs)"},
                ],
                "trend": trend,
                "fetched_at": time.strftime("%Y-%m-%d %H:%M:%S"),
            }))
        except Exception as exc:
            return jsonify(json_safe({
                "error": f"Superpower DHIS2 query failed: {str(exc)}"
            })), 500

    # ── DHIS2 Live Query: Newly Started on ART ───────────────────────────
    @app.get("/api/dhis/query-art")
    def dhis_query_art() -> object:
        """Query Newly Started on ART (Tx_New) from live DHIS2 only.
        ALWAYS fetches LAST_6_MONTHS for trend data; filters on frontend side.
        """
        facility = (request.args.get("facility") or "").strip()
        month = (request.args.get("month") or "").strip()
        county = (request.args.get("county") or "").strip()
        subcounty = (request.args.get("subCounty") or "").strip()

        DX_TX_NEW = "gv7bbGesTTJ"
        dhis_base = os.getenv("DHIS_BASE_URL") or "http://ereporting.chak.or.ke:8500/api/"
        username = os.getenv("DHIS_USERNAME") or "Johnbrian"
        password = os.getenv("DHIS_PASSWORD") or "JOHNb123\\"

        all_rows = []
        total = 0
        total_requested_month = 0

        # ── Helper: resolve org unit name → ID ──────────────────────────
        # Known county UIDs
        KNOWN_COUNTIES = {
            "meru": "Y52XNJ50hYb",
            "kiambu": "uYVuOOi3Sdo",
            "nairobi": "HjfgtUvCQmM",
            "mombasa": "YqRHObuZC00",
            "kisumu": "HhOLwQKSAfM",
            "makueni": "Q0Fqz15LkGk",
            "kitui": "COSuCgBjjM3",
            "machakos": "HfLKw6JcHdv",
            "kilifi": "NG7ZEAV0hy4",
        }

        def resolve_ou(name, search_type="facility"):
            """search_type: 'facility' or 'county'"""
            name_lower = name.lower().strip()
            # For county lookups, use known counties mapping directly
            if search_type == "county":
                for county_name, uid in KNOWN_COUNTIES.items():
                    if name_lower == county_name or name_lower.startswith(county_name):
                        return uid
                return None
            # For facility lookups, use the facilities dictionary
            try:
                _, df_fac = _superpower_load_dict()
                m = df_fac[df_fac["name"].str.lower().str.contains(name_lower, na=False)]
                if not m.empty:
                    fac = m[m["level"].astype(str) == "5"]
                    if not fac.empty:
                        return fac.iloc[0]["id"]
                    return m.iloc[0]["id"]
            except Exception:
                pass
            return None

        # ── Source 1: Live DHIS2 ─────────────────────────────────────────
        from requests.auth import HTTPBasicAuth
        import requests as _req

        auth = HTTPBasicAuth(username, password)
        url_base = dhis_base.rstrip("/") + "/analytics.json"

        ou_id = None
        ou_label = "All Facilities"
        if facility and facility.lower() != "all":
            ou_id = resolve_ou(facility, "facility") or "uwvOG2N2Cmt"
            ou_label = facility
        elif county and county.lower() != "all":
            ou_id = resolve_ou(county, "county") or "Y52XNJ50hYb"
            ou_label = county

        # Always fetch LAST_6_MONTHS for rich trend data
        pe = "LAST_6_MONTHS"
        month_clean = ""
        if month and month.lower() != "all":
            month_clean = re.sub(r"[^0-9]", "", month)[:6]

        params = {"dimension": [f"dx:{DX_TX_NEW}", f"pe:{pe}"]}
        if ou_id:
            params["dimension"].append(f"ou:{ou_id}")

        try:
            resp = _req.get(url_base, params=params, auth=auth, timeout=120)
            if resp.ok:
                dhis_data = resp.json()
                # Map column names from response headers
                headers = [h.get("name", "").lower() for h in dhis_data.get("headers", [])]
                col_map = {name: i for i, name in enumerate(headers)}
                dx_idx = col_map.get("dx", 0)
                ou_idx = col_map.get("ou")  # None if ou not in response
                pe_idx = col_map.get("pe", 2)
                val_idx = col_map.get("value", 3)

                meta = dhis_data.get("metaData", {})
                items_map = {k: v.get("name", k) for k, v in meta.get("items", {}).items()}

                for row in dhis_data.get("rows", []):
                    pe_str = row[pe_idx] if len(row) > pe_idx and row[pe_idx] else pe
                    val = float(row[val_idx]) if len(row) > val_idx and row[val_idx] else 0
                    ou_code = ""
                    ou_name = ou_label or "All"
                    if ou_idx is not None and len(row) > ou_idx and row[ou_idx]:
                        ou_code = row[ou_idx]
                        ou_name = items_map.get(ou_code, ou_code)
                    pe_label = items_map.get(pe_str, pe_str) if pe_str else pe_str
                    row_data = {
                        "period": pe_str,
                        "period_label": pe_label,
                        "org_unit": ou_name,
                        "org_unit_id": ou_code,
                        "facility": ou_name,
                        "value": val,
                        "source": "live_dhis2",
                    }
                    all_rows.append(row_data)
                    total += val
                    # Track total for the specifically requested month
                    if month_clean and pe_str == month_clean:
                        total_requested_month += val
        except Exception as exc:
            return jsonify(json_safe({"ok": False, "error": f"DHIS2 query failed: {str(exc)}"}))

        # Sort by period
        all_rows.sort(key=lambda r: str(r.get("period", "")))

        return jsonify(json_safe({
            "ok": True,
            "indicator": "Newly Started on ART (Tx_New STA)",
            "total": total,
            "total_requested_month": total_requested_month if month_clean else total,
            "requested_month": month_clean,
            "rows": all_rows,
            "row_count": len(all_rows),
            "org_unit": ou_id or "all",
            "org_unit_label": ou_label,
            "period": pe,
        }))

    # ── Generic DHIS2 query (any indicator) ───────────────────────────────
    @app.get("/api/dhis/query")
    def dhis_query() -> object:
        """Generic DHIS2 query — pass ?q=your question."""
        question = (request.args.get("q") or "").strip()
        if not question:
            return jsonify(json_safe({"ok": False, "error": "Missing 'q' parameter"}))

        if not HAS_SUPERPOWER:
            return jsonify(json_safe({"ok": False, "error": "Superpower module not available"}))

        try:
            url = _superpower_generate_url(question)
            if not url:
                return jsonify(json_safe({"ok": False, "error": "No URL generated", "question": question}))
            result = _superpower_fetch_result(url, user_question=question)
            return jsonify(json_safe({
                "ok": result.get("ok", False),
                "question": question,
                "url": url,
                "total": result.get("total", 0),
                "rows": result.get("rows", []),
                "error": result.get("error"),
                "message": result.get("message"),
            }))
        except Exception as exc:
            return jsonify(json_safe({"ok": False, "error": str(exc)})), 500

    # ── TX_CURR Analytics: Gender Breakdown ────────────────────────────
    @app.get("/api/hiv-treatment/tx-curr-gender")
    def tx_curr_gender() -> object:
        """TX_CURR by gender for the selected location.
        ?county=...&subcounty=...&facility=...&period=LAST_12_MONTHS
        Returns {period_series: [...], latest_gender: {Male: N, Female: N}}
        """
        county = (request.args.get("county") or "Meru County").strip()
        sc_filter = request.args.get("subcounty", "").strip()
        fac_filter = request.args.get("facility", "").strip()
        pe = (request.args.get("period") or "LAST_12_MONTHS").strip()
        ou_id, _ = _resolve_ou_ids(county, sc_filter or None, fac_filter or None)
        TX_CURR_DX = INDICATOR_SPECS["tx_curr"]["aggregate"]

        try:
            raw = _dhis2_fetch(TX_CURR_DX, ou_id, pe, None)
        except Exception as exc:
            return jsonify(json_safe({"ok": False, "error": str(exc)})), 500
        return jsonify(json_safe({"ok": True, "data": raw}))

    # ── TX_CURR Analytics: By Finer Age Group ─────────────────────────
    @app.get("/api/hiv-treatment/tx-curr-age")
    def tx_curr_age() -> object:
        """TX_CURR by finer age group for the selected location.
        ?county=...&subcounty=...&facility=...&pe=202601 (single month)
        Returns {period, age_data: [{age, value}]}
        """
        county = (request.args.get("county") or "Meru County").strip()
        sc_filter = request.args.get("subcounty", "").strip()
        fac_filter = request.args.get("facility", "").strip()
        pe = (request.args.get("pe") or "202605").strip()
        ou_id, _ = _resolve_ou_ids(county, sc_filter or None, fac_filter or None)
        TX_CURR_DX = INDICATOR_SPECS["tx_curr"]["aggregate"]
        AGE_DIM = "PiDJ9GbMZ0B"

        try:
            raw = _dhis2_fetch(TX_CURR_DX, ou_id, pe, None)
        except Exception as exc:
            return jsonify(json_safe({"ok": False, "error": str(exc)})), 500
        return jsonify(json_safe({"ok": True, "data": raw, "period": pe}))

    # ── TX_CURR Analytics: Yearly Trends ──────────────────────────────
    @app.get("/api/hiv-treatment/tx-curr-yearly")
    def tx_curr_yearly() -> object:
        """TX_CURR yearly totals for the selected location.
        ?county=...&subcounty=...&facility=...&pe=2023;2024;2025;2026
        Returns {data: {year_label: value}}
        """
        county = (request.args.get("county") or "Meru County").strip()
        sc_filter = request.args.get("subcounty", "").strip()
        fac_filter = request.args.get("facility", "").strip()
        pe = (request.args.get("pe") or "2023;2024;2025;2026").strip()
        ou_id, _ = _resolve_ou_ids(county, sc_filter or None, fac_filter or None)
        TX_CURR_DX = INDICATOR_SPECS["tx_curr"]["aggregate"]

        try:
            raw = _dhis2_fetch(TX_CURR_DX, ou_id, pe, None)
        except Exception as exc:
            return jsonify(json_safe({"ok": False, "error": str(exc)})), 500
        return jsonify(json_safe({"ok": True, "data": raw}))

    # ── TX_CURR Analytics: MMD Breakdown ──────────────────────────────
    @app.get("/api/hiv-treatment/tx-curr-mmd")
    def tx_curr_mmd() -> object:
        """TX_CURR MMD regimen breakdown for the selected location.
        ?county=...&subcounty=...&facility=...&period=LAST_12_MONTHS
        Returns {period_series: {...}, latest_data: {regimen_label: value}}
        """
        county = (request.args.get("county") or "Meru County").strip()
        sc_filter = request.args.get("subcounty", "").strip()
        fac_filter = request.args.get("facility", "").strip()
        pe = (request.args.get("period") or "LAST_12_MONTHS").strip()
        ou_id, _ = _resolve_ou_ids(county, sc_filter or None, fac_filter or None)

        MMD_DX = (
            "TNAf1ystLF3;JOldQxWZWso;HzXPYZqLgqj;g8mOybcTwmL;"
            "EgNQnR23En1;KsDSjjJo6GD;VIz7xRli13H;KEAYcGVL6Bk;qDjo1L1VfmP"
        )

        try:
            raw = _dhis2_fetch(MMD_DX, ou_id, pe, None)
        except Exception as exc:
            return jsonify(json_safe({"ok": False, "error": str(exc)})), 500
        return jsonify(json_safe({"ok": True, "data": raw}))

    # ── TX_CURR Analytics: Month-on-Month Change ──────────────────────
    @app.get("/api/hiv-treatment/tx-curr-mom")
    def tx_curr_mom() -> object:
        """TX_CURR month-on-month change for the selected location.
        ?county=...&subcounty=...&facility=...&period=LAST_12_MONTHS
        Returns {changes: [{period, current, previous, change, change_pct}]}
        """
        county = (request.args.get("county") or "Meru County").strip()
        sc_filter = request.args.get("subcounty", "").strip()
        fac_filter = request.args.get("facility", "").strip()
        pe = (request.args.get("period") or "LAST_12_MONTHS").strip()
        ou_id, _ = _resolve_ou_ids(county, sc_filter or None, fac_filter or None)
        TX_CURR_DX = INDICATOR_SPECS["tx_curr"]["aggregate"]

        try:
            raw = _dhis2_fetch(TX_CURR_DX, ou_id, pe, None)
        except Exception as exc:
            return jsonify(json_safe({"ok": False, "error": str(exc)})), 500

        # Compute MoM changes
        import re
        from datetime import datetime

        def _parse_period(label):
            m = re.match(r"(\w+)\s+(\d{4})", label)
            if m:
                return datetime.strptime(f"{m.group(1)} {m.group(2)}", "%B %Y")
            return datetime.min

        sorted_periods = sorted(raw.items(), key=lambda x: _parse_period(x[0]))
        changes = []
        for i in range(1, len(sorted_periods)):
            prev_label, prev_val = sorted_periods[i-1]
            curr_label, curr_val = sorted_periods[i]
            change = curr_val - prev_val
            change_pct = round((change / prev_val * 100), 1) if prev_val else 0
            changes.append({
                "period": curr_label,
                "current": curr_val,
                "previous": prev_val,
                "change": change,
                "change_pct": change_pct,
            })

        return jsonify(json_safe({"ok": True, "changes": changes}))

    # ── JTP Regimen Distribution (like DHIS2 TX_Curr Regimens) ──────
    @app.get("/api/hiv-treatment/jtp-regimens")
    def jtp_regimens() -> object:
        """Fetches JTP regimen data for donut chart.
        ?county=...&subcounty=...&facility=...&period=LAST_12_MONTHS
        Uses ART Optimization DX IDs: 1st Line, 2nd Line, 3rd Line, DTG
        Returns {ok, regimens: [{label, id, value}], latest_period}
        """
        county = (request.args.get("county") or "Meru County").strip()
        sc_filter = request.args.get("subcounty", "").strip()
        fac_filter = request.args.get("facility", "").strip()
        pe = (request.args.get("period") or "LAST_12_MONTHS").strip()
        ou_id, _ = _resolve_ou_ids(county, sc_filter or None, fac_filter or None)

        REGIMEN_DX = {
            "1st Line ART": "zZGNba5d34c",
            "2nd Line ART": "F0xtjHxDZ2e",
            "3rd Line ART": "Pk1PMmG4ml7",
            "On DTG": "s62uidROGjG",
            "Eligible DTG": "bsQdHW8sJ4b",
            "Active on EFV-600": "lr1YorhNrJT",
            "Active on EFV-400": "ggO3YzjB9j4",
            "Active on PI": "Z4g3jskQn9c",
            "Viremia Clinic": "JGIZOGP6bGU",
        }
        dx_all = ";".join(REGIMEN_DX.values())

        try:
            raw = _dhis2_fetch(dx_all, ou_id, pe, None)
        except Exception as exc:
            return jsonify(json_safe({"ok": False, "error": str(exc)})), 500

        # Group by period
        from collections import defaultdict
        period_data: dict[str, dict[str, float]] = defaultdict(dict)
        # raw is {"Period Label": total_value} with all regimens summed together
        # Need to fetch separately per DX
        regimens = []
        latest_period = ""
        try:
            from concurrent.futures import ThreadPoolExecutor, as_completed
            with ThreadPoolExecutor(max_workers=6) as executor:
                fut_map = {}
                for label, dx_id in REGIMEN_DX.items():
                    fut = executor.submit(_dhis2_fetch, dx_id, ou_id, pe, None)
                    fut_map[fut] = label
                for fut in as_completed(fut_map):
                    label = fut_map[fut]
                    try:
                        data = fut.result()
                    except Exception:
                        data = {}
                    # Get latest non-zero value
                    sorted_periods = sorted(data.keys(), key=_period_sort_key)
                    val = 0
                    lp = ""
                    for p in reversed(sorted_periods):
                        v = round(float(data.get(p, 0)), 1)
                        if v > 0:
                            val = v
                            lp = p
                            break
                    regimens.append({"label": label, "id": REGIMEN_DX[label], "value": val, "period": lp})
                    if lp and lp > latest_period:
                        latest_period = lp
        except Exception as exc:
            return jsonify(json_safe({"ok": False, "error": str(exc)})), 500

        # Calculate total
        total = sum(r["value"] for r in regimens)
        return jsonify(json_safe({
            "ok": True,
            "regimens": regimens,
            "total": total,
            "latest_period": latest_period,
        }))

    # ── TX_CURR by Gender (Male/Female aggregate from DX elements) ──
    @app.get("/api/hiv-treatment/tx-curr-gender-split")
    def tx_curr_gender_split() -> object:
        """TX_CURR Male vs Female from gender-specific DX elements.
        ?county=...&subcounty=...&facility=...&period=LAST_12_MONTHS
        Returns {ok, male: N, female: N, latest_period, trend: [{period, male, female}]}
        """
        county = (request.args.get("county") or "Meru County").strip()
        sc_filter = request.args.get("subcounty", "").strip()
        fac_filter = request.args.get("facility", "").strip()
        pe = (request.args.get("period") or "LAST_12_MONTHS").strip()
        ou_id, _ = _resolve_ou_ids(county, sc_filter or None, fac_filter or None)

        # Female TX_Curr STA DX IDs (all ages)
        FEMALE_DX = [
            "iaa4KseNcet",  # TX_Curr STA <1,F
            "hJmFsJUytKD",  # TX_Curr STA 1-4,F
            "FwRKImEnyEs",  # TX_Curr STA 5-9,F
            "Aiq7hJDqUEe",  # TX_Curr STA 10-14,F
            "n5ySsHEkFrs",  # TX_Curr STA 15-19,F
            "qo1sG5nv3sM",  # TX_Curr STA 20-24,F
            "J1djCE9rcZZ",  # TX_Curr STA 25-29,F
            "SHzQklQSFti",  # TX_Curr STA 30-34,F
            "brMgg890UfA",  # TX_Curr STA 35-39,F
            "vEOthZE5MwG",  # TX_Curr STA 40-44,F
            "rL9iyqtuW5w",  # TX_Curr STA 45-49,F
            "FkNNEFbIWiM",  # TX_Curr STA 50-54,F
            "NEb6Ty89bbF",  # TX_Curr STA 55-59,F
            "MMFZc5KvI8m",  # TX_Curr STA 60-64,F
            "bLbb816Lep0",  # TX_Curr STA 65+,F
        ]
        MALE_DX = [
            "Q8ErsVgUUy7",  # TX_Curr STA <1,M
            "P8UoaFZ9whV",  # TX_Curr STA 1-4,M
            "CBoJcoKZ7Iy",  # TX_Curr STA 5-9,M
            "UoCnviagVgb",  # TX_Curr STA 10-14,M
            "LofgXYRFD02",  # TX_Curr STA 15-19,M
            "TKSQgnyBukU",  # TX_Curr STA 20-24,M
            "WwOocFBoNQj",  # TX_Curr STA 25-29,M
            "g9yMnhmPQ58",  # TX_Curr STA 30-34,M
            "isy3s3kUVQC",  # TX_Curr STA 35-39,M
            "xpZhQHWpqL8",  # TX_Curr STA 40-44,M
            "F4ZrXG2G3Kv",  # TX_Curr STA 45-49,M
            "dWDqkhd9IAv",  # TX_Curr STA 50-54,M
            "N01LC1ThJUT",  # TX_Curr STA 55-59,M
            "o9YAn2dQuXx",  # TX_Curr STA 60-64,M
            "Ex31nkiTRuJ",  # TX_Curr STA 65+,M
        ]
        female_dx_all = ";".join(FEMALE_DX)
        male_dx_all = ";".join(MALE_DX)

        try:
            from concurrent.futures import ThreadPoolExecutor, as_completed
            results = {}
            with ThreadPoolExecutor(max_workers=2) as executor:
                f_fut = executor.submit(_dhis2_fetch, female_dx_all, ou_id, pe, None)
                m_fut = executor.submit(_dhis2_fetch, male_dx_all, ou_id, pe, None)
                for fut in as_completed([f_fut, m_fut]):
                    try:
                        data = fut.result()
                    except Exception:
                        data = {}
                    if fut == f_fut:
                        results["female"] = data
                    else:
                        results["male"] = data
        except Exception as exc:
            return jsonify(json_safe({"ok": False, "error": str(exc)})), 500

        female_data = results.get("female", {})
        male_data = results.get("male", {})

        # Build trend
        all_periods = sorted(set(list(female_data.keys()) + list(male_data.keys())), key=_period_sort_key)
        trend = []
        for p in all_periods:
            f_val = round(float(female_data.get(p, 0)), 1)
            m_val = round(float(male_data.get(p, 0)), 1)
            trend.append({"period": p, "label": p, "female": f_val, "male": m_val})

        # Latest values
        latest_f = round(float(female_data.get(all_periods[-1], 0)), 1) if all_periods else 0
        latest_m = round(float(male_data.get(all_periods[-1], 0)), 1) if all_periods else 0

        return jsonify(json_safe({
            "ok": True,
            "male": latest_m,
            "female": latest_f,
            "total": latest_f + latest_m,
            "latest_period": all_periods[-1] if all_periods else "",
            "trend": trend,
        }))

    # ── TX_CURR by Finer Age-Group (using age-specific DX elements) ──
    @app.get("/api/hiv-treatment/tx-curr-age-split")
    def tx_curr_age_split() -> object:
        """TX_CURR by finer age groups using age-specific DX elements.
        ?county=...&subcounty=...&facility=...&period=LAST_12_MONTHS
        Returns {ok, age_data: [{age, value}], latest_period, trend}
        """
        county = (request.args.get("county") or "Meru County").strip()
        sc_filter = request.args.get("subcounty", "").strip()
        fac_filter = request.args.get("facility", "").strip()
        pe = (request.args.get("period") or "LAST_12_MONTHS").strip()
        ou_id, _ = _resolve_ou_ids(county, sc_filter or None, fac_filter or None)

        AGE_DX_MAP = {
            "<1": "iaa4KseNcet;Q8ErsVgUUy7",
            "1-4": "hJmFsJUytKD;P8UoaFZ9whV",
            "5-9": "FwRKImEnyEs;CBoJcoKZ7Iy",
            "10-14": "Aiq7hJDqUEe;UoCnviagVgb",
            "15-19": "n5ySsHEkFrs;LofgXYRFD02",
            "20-24": "qo1sG5nv3sM;TKSQgnyBukU",
            "25-29": "J1djCE9rcZZ;WwOocFBoNQj",
            "30-34": "SHzQklQSFti;g9yMnhmPQ58",
            "35-39": "brMgg890UfA;isy3s3kUVQC",
            "40-44": "vEOthZE5MwG;xpZhQHWpqL8",
            "45-49": "rL9iyqtuW5w;F4ZrXG2G3Kv",
            "50-54": "FkNNEFbIWiM;dWDqkhd9IAv",
            "55-59": "NEb6Ty89bbF;N01LC1ThJUT",
            "60-64": "MMFZc5KvI8m;o9YAn2dQuXx",
            "65+": "bLbb816Lep0;Ex31nkiTRuJ",
        }

        try:
            from concurrent.futures import ThreadPoolExecutor, as_completed
            age_results = {}
            with ThreadPoolExecutor(max_workers=8) as executor:
                fut_map = {}
                for age_label, dx_ids in AGE_DX_MAP.items():
                    fut = executor.submit(_dhis2_fetch, dx_ids, ou_id, pe, None)
                    fut_map[fut] = age_label
                for fut in as_completed(fut_map):
                    age_label = fut_map[fut]
                    try:
                        data = fut.result()
                    except Exception:
                        data = {}
                    age_results[age_label] = data
        except Exception as exc:
            return jsonify(json_safe({"ok": False, "error": str(exc)})), 500

        # Build age_data (latest period values)
        all_periods = set()
        for age_data in age_results.values():
            all_periods.update(age_data.keys())
        sorted_periods = sorted(all_periods, key=_period_sort_key)
        latest_period = sorted_periods[-1] if sorted_periods else ""

        age_data = []
        AGE_ORDER = ["<1","1-4","5-9","10-14","15-19","20-24","25-29","30-34","35-39","40-44","45-49","50-54","55-59","60-64","65+"]
        for age in AGE_ORDER:
            d = age_results.get(age, {})
            val = round(float(d.get(latest_period, 0)), 1) if latest_period else 0
            if val > 0:
                age_data.append({"age": age, "value": val})

        # Build trend (total TX_CURR across all ages)
        trend = []
        for p in sorted_periods:
            total = sum(round(float(age_results.get(a, {}).get(p, 0)), 1) for a in AGE_ORDER)
            trend.append({"period": p, "label": p, "value": total})

        return jsonify(json_safe({
            "ok": True,
            "age_data": age_data,
            "latest_period": latest_period,
            "trend": trend,
        }))

    # ── TX_NEW by Gender (Male/Female aggregate) ─────────────────────
    @app.get("/api/hiv-treatment/tx-new-gender-split")
    def tx_new_gender_split() -> object:
        county = (request.args.get("county") or "Meru County").strip()
        sc_filter = request.args.get("subcounty", "").strip()
        fac_filter = request.args.get("facility", "").strip()
        pe = (request.args.get("period") or "LAST_12_MONTHS").strip()
        ou_id, _ = _resolve_ou_ids(county, sc_filter or None, fac_filter or None)

        FEMALE_DX = [
            "X7QikQUsYB1", "MOqDhGiw7W6", "BFbmB3WxGPd", "xz2f0oONxQx",
            "f59E1kimKqe", "UjIzCVxESAz", "Y4jKOMblgII", "hUHq4KO9YMz",
            "tWTgIibsKJ5", "AFdikiNpC3e", "GMoRCzegC6C", "Bs5etPcLz7w",
            "Ps8a7Mv1xIn", "sQcd8UD8Mrs",
        ]
        MALE_DX = [
            "s9iBEnfSHhh", "JprDjnAyB0f", "PG5Ynz9xGCu", "SNOcc1Tq2iH",
            "VqYNMLji5U5", "wJWCrZVh1iu", "jjXJNig8fxs", "Lbs5RUpnwPD",
            "cQrYHDWkY2y", "QRV2YRNGYJ6", "Mt9G8jCODUw", "FDRjPKGGVC9",
            "ShO7o3bHsNr", "ivLPgJtKgcN",
        ]

        try:
            from concurrent.futures import ThreadPoolExecutor, as_completed
            results = {}
            with ThreadPoolExecutor(max_workers=2) as executor:
                f_fut = executor.submit(_dhis2_fetch, ";".join(FEMALE_DX), ou_id, pe, None)
                m_fut = executor.submit(_dhis2_fetch, ";".join(MALE_DX), ou_id, pe, None)
                for fut in as_completed([f_fut, m_fut]):
                    try:
                        data = fut.result()
                    except Exception:
                        data = {}
                    results["female" if fut == f_fut else "male"] = data
        except Exception as exc:
            return jsonify(json_safe({"ok": False, "error": str(exc)})), 500

        female_data = results.get("female", {})
        male_data = results.get("male", {})
        all_periods = sorted(set(list(female_data.keys()) + list(male_data.keys())), key=_period_sort_key)
        trend = []
        for p in all_periods:
            f_val = round(float(female_data.get(p, 0)), 1)
            m_val = round(float(male_data.get(p, 0)), 1)
            trend.append({"period": p, "label": p, "female": f_val, "male": m_val})
        latest_f = round(float(female_data.get(all_periods[-1], 0)), 1) if all_periods else 0
        latest_m = round(float(male_data.get(all_periods[-1], 0)), 1) if all_periods else 0
        return jsonify(json_safe({
            "ok": True, "male": latest_m, "female": latest_f,
            "total": latest_f + latest_m,
            "latest_period": all_periods[-1] if all_periods else "",
            "trend": trend,
        }))

    # ── TX_NEW by Finer Age-Group ────────────────────────────────────
    @app.get("/api/hiv-treatment/tx-new-age-split")
    def tx_new_age_split() -> object:
        county = (request.args.get("county") or "Meru County").strip()
        sc_filter = request.args.get("subcounty", "").strip()
        fac_filter = request.args.get("facility", "").strip()
        pe = (request.args.get("period") or "LAST_12_MONTHS").strip()
        ou_id, _ = _resolve_ou_ids(county, sc_filter or None, fac_filter or None)

        AGE_DX_MAP = {
            "1-4": "X7QikQUsYB1;s9iBEnfSHhh",
            "5-9": "MOqDhGiw7W6;JprDjnAyB0f",
            "10-14": "BFbmB3WxGPd;PG5Ynz9xGCu",
            "15-19": "xz2f0oONxQx;SNOcc1Tq2iH",
            "20-24": "f59E1kimKqe;VqYNMLji5U5",
            "25-29": "UjIzCVxESAz;wJWCrZVh1iu",
            "30-34": "Y4jKOMblgII;jjXJNig8fxs",
            "35-39": "hUHq4KO9YMz;Lbs5RUpnwPD",
            "40-44": "tWTgIibsKJ5;cQrYHDWkY2y",
            "45-49": "AFdikiNpC3e;QRV2YRNGYJ6",
            "50-54": "GMoRCzegC6C;Mt9G8jCODUw",
            "55-59": "Bs5etPcLz7w;FDRjPKGGVC9",
            "60-64": "Ps8a7Mv1xIn;ShO7o3bHsNr",
            "65+": "sQcd8UD8Mrs;ivLPgJtKgcN",
        }
        AGE_ORDER = ["1-4","5-9","10-14","15-19","20-24","25-29","30-34","35-39","40-44","45-49","50-54","55-59","60-64","65+"]

        try:
            from concurrent.futures import ThreadPoolExecutor, as_completed
            age_results = {}
            with ThreadPoolExecutor(max_workers=8) as executor:
                fut_map = {}
                for age_label, dx_ids in AGE_DX_MAP.items():
                    fut = executor.submit(_dhis2_fetch, dx_ids, ou_id, pe, None)
                    fut_map[fut] = age_label
                for fut in as_completed(fut_map):
                    age_label = fut_map[fut]
                    try:
                        data = fut.result()
                    except Exception:
                        data = {}
                    age_results[age_label] = data
        except Exception as exc:
            return jsonify(json_safe({"ok": False, "error": str(exc)})), 500

        all_periods = set()
        for age_data in age_results.values():
            all_periods.update(age_data.keys())
        sorted_periods = sorted(all_periods, key=_period_sort_key)
        latest_period = sorted_periods[-1] if sorted_periods else ""
        age_data = []
        for age in AGE_ORDER:
            d = age_results.get(age, {})
            val = round(float(d.get(latest_period, 0)), 1) if latest_period else 0
            if val > 0:
                age_data.append({"age": age, "value": val})
        return jsonify(json_safe({
            "ok": True, "age_data": age_data,
            "latest_period": latest_period,
        }))

    # ── Homepage Summary Endpoint ─────────────────────────────────────
    @app.get("/api/homepage/summary")
    def homepage_summary() -> object:
        """Consolidated DHIS2 summary for the homepage dashboard.
        Returns TX_CURR, TX_NEW, and HTS uptake data in one response.
        Params: ?county=...&subcounty=...&period=LAST_12_MONTHS
        """
        county = (request.args.get("county") or "Meru County").strip()
        pe = (request.args.get("period") or "LAST_12_MONTHS").strip()
        sc_filter = request.args.get("subcounty", "").strip()

        ou_id, _ = _resolve_ou_ids(county, sc_filter or None, None)

        from concurrent.futures import ThreadPoolExecutor, as_completed

        errors = []
        result = {}

        # ── DX IDs ──
        TX_NEW_DX = INDICATOR_SPECS["tx_new"]["aggregate"]    # vTTEybkXZ53
        TX_CURR_DX = INDICATOR_SPECS["tx_curr"]["aggregate"]  # kgzd9LfXZXq

        # HTS entry-point DXs for tested count
        HTS_TESTED_DX = [
            "ymKviaHZtQN","vFlUDposW0Y","XKAlilawdhN","THJbtDzxplR",
            "Lwtqyjus0Mb","QBsyLQZRdiH","XYhYAMivUX5","J4zibSjbBCt",
        ]
        HTS_POSITIVE_DX = "CcOr3MB7Mh4"

        # ── Parallel fetch all 4 DX groups ──
        fetch_tasks = {
            "tx_new": TX_NEW_DX,
            "tx_curr": TX_CURR_DX,
            "hts_tested": HTS_TESTED_DX,
            "hts_positive": HTS_POSITIVE_DX,
        }

        fetched = {}
        with ThreadPoolExecutor(max_workers=4) as executor:
            futures = {}
            for key, dx in fetch_tasks.items():
                futures[executor.submit(_dhis2_fetch, dx, ou_id, pe, None)] = key
            for future in as_completed(futures):
                key = futures[future]
                try:
                    fetched[key] = future.result()
                except Exception as exc:
                    errors.append(f"{key}: {exc}")
                    fetched[key] = {}

        # ── Build trend ──
        all_periods = sorted(
            set().union(*[d.keys() for d in fetched.values()]),
            key=_period_sort_key,
        )

        def _label(p):
            return f"{p[:4]}-{p[4:]}" if len(str(p)) == 6 else str(p)

        tx_new_trend = []
        tx_curr_trend = []
        hts_trend = []

        for p in all_periods:
            tx_new_trend.append({
                "period": p, "label": _label(p),
                "value": round(float(fetched["tx_new"].get(p, 0)), 1),
            })
            tx_curr_trend.append({
                "period": p, "label": _label(p),
                "value": round(float(fetched["tx_curr"].get(p, 0)), 1),
            })
            tested = float(fetched["hts_tested"].get(p, 0))
            positive = float(fetched["hts_positive"].get(p, 0))
            hts_trend.append({
                "period": p, "label": _label(p),
                "tested": round(tested, 1),
                "positive": round(positive, 1),
                "positivity_rate": round((positive / tested * 100) if tested > 0 else 0, 1),
            })

        # Latest month KPIs
        latest = all_periods[-1] if all_periods else None
        kpis = {}
        if latest:
            kpis = {
                "label": _label(latest),
                "tx_curr": round(float(fetched["tx_curr"].get(latest, 0)), 1),
                "tx_new": round(float(fetched["tx_new"].get(latest, 0)), 1),
                "hts_tested": round(float(fetched["hts_tested"].get(latest, 0)), 1),
                "hts_positive": round(float(fetched["hts_positive"].get(latest, 0)), 1),
            }
            kpis["positivity_rate"] = round(
                (kpis["hts_positive"] / kpis["hts_tested"] * 100)
                if kpis["hts_tested"] > 0 else 0, 1
            )

        return jsonify(json_safe({
            "source": "dhis2_live",
            "county": county,
            "subcounty": sc_filter or None,
            "ou_id": ou_id,
            "period_range": pe,
            "latest": kpis,
            "tx_new_trend": tx_new_trend,
            "tx_curr_trend": tx_curr_trend,
            "hts_trend": hts_trend,
            "fetched_at": time.strftime("%Y-%m-%d %H:%M:%S"),
            "errors": errors if errors else None,
        }))

    # ── Generic: Search DHIS2 Data Elements by name ──────────────────
    @app.get("/api/dhis2/search-elements")
    def search_dhis2_elements() -> object:
        """Search master_data_elements.csv by name/pattern.
        Params: ?q=tx_pvls&limit=20
        Returns list of matching {id, name, displayName, shortName}
        """
        q = (request.args.get("q") or "").strip().lower()
        limit = int(request.args.get("limit", "20"))

        if not q:
            return jsonify(json_safe({"elements": [], "query": q, "total": 0}))

        dict_path = BASE_DIR.parent / "dictionaries" / "master_data_elements.csv"
        if not dict_path.exists():
            return jsonify(json_safe({"error": "Dictionary not found"})), 500

        try:
            df = pd.read_csv(dict_path)
        except Exception as exc:
            return jsonify(json_safe({"error": str(exc)})), 500

        # Search across name, displayName, shortName, id
        mask = (
            df["name"].astype(str).str.lower().str.contains(q, na=False) |
            df["displayName"].astype(str).str.lower().str.contains(q, na=False) |
            df["shortName"].astype(str).str.lower().str.contains(q, na=False) |
            df["id"].astype(str).str.lower().str.contains(q, na=False)
        )
        results = df[mask].head(limit * 3)  # grab extra for dedup
        elements = []
        seen = set()
        for _, row in results.iterrows():
            uid = str(row["id"])
            if uid in seen:
                continue
            seen.add(uid)
            elements.append({
                "id": uid,
                "name": str(row.get("name", "")),
                "displayName": str(row.get("displayName", "")),
                "shortName": str(row.get("shortName", "")),
            })
            if len(elements) >= limit:
                break

        return jsonify(json_safe({
            "elements": elements,
            "query": q,
            "total": len(elements),
        }))

    # ── Generic: Query ANY data element from DHIS2 live ──────────────
    @app.get("/api/dhis2/query")
    def generic_dhis2_query() -> object:
        """Query any DHIS2 data element(s) live.
        Params: ?dx=JGd3MwmKBuM&county=Meru+County
                &subcounty=...&facility=...&period=LAST_12_MONTHS
        dx can be a single UID or comma-separated list.
        Returns monthly trend data.
        """
        dx_str = (request.args.get("dx") or "").strip()
        county = (request.args.get("county") or "Meru County").strip()
        pe = (request.args.get("period") or "LAST_12_MONTHS").strip()
        sc_filter = request.args.get("subcounty", "").strip()
        fac_filter = request.args.get("facility", "").strip()

        if not dx_str:
            return jsonify(json_safe({"error": "dx parameter is required"})), 400

        dx_ids = [d.strip() for d in dx_str.split(",") if d.strip()]
        ou_id, _ = _resolve_ou_ids(county, sc_filter or None, fac_filter or None)

        raw = _dhis2_fetch(dx_ids, ou_id, pe, None)

        # Resolve names from master dictionary
        name_map = {}
        dict_path = BASE_DIR.parent / "dictionaries" / "master_data_elements.csv"
        if dict_path.exists():
            try:
                df_dict = pd.read_csv(dict_path)
                for _, r in df_dict.iterrows():
                    name_map[str(r["id"])] = str(r.get("name", ""))
            except Exception:
                pass

        trend = []
        all_periods = sorted(set(raw.keys()), key=_period_sort_key)
        for p in all_periods:
            trend.append({
                "period": p,
                "label": f"{p[:4]}-{p[4:]}" if len(str(p)) == 6 else str(p),
                "value": round(float(raw.get(p, 0)), 1),
            })

        return jsonify(json_safe({
            "dx": dx_ids,
            "dx_names": {d: name_map.get(d, "") for d in dx_ids},
            "source": "dhis2_live",
            "county": county,
            "subcounty": sc_filter or None,
            "facility": fac_filter or None,
            "ou_id": ou_id,
            "period_range": pe,
            "total": round(sum(float(raw.get(p, 0)) for p in all_periods), 1),
            "trend": trend,
            "fetched_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        }))

    # ── Register PBIX Dashboard Blueprint (16 CHAK Visuals pages) ──
    try:
        from pbix_dashboards import register_pbix_blueprint
        register_pbix_blueprint(app, _dhis2_fetch, _resolve_ou_ids)
        print("[PBIX] Dashboard blueprint registered successfully")
    except ImportError as exc:
        print(f"[PBIX] Could not load pbix_dashboards module: {exc}")
    except Exception as exc:
        print(f"[PBIX] Error registering PBIX blueprint: {exc}")

    # ────────────────────────────────────────────────────────────
    # PROJECT PERFORMANCE MONITORING — API Endpoints
    # ────────────────────────────────────────────────────────────

    @app.get("/api/project-portfolio")
    def project_portfolio() -> object:
        """Return the full project portfolio dashboard."""
        try:
            data = load_project_performance_data()
            if "error" in data:
                return jsonify(json_safe({"ok": False, "error": data["error"]}))
            return jsonify(json_safe({"ok": True, **data}))
        except Exception as exc:
            return jsonify(json_safe({"ok": False, "error": str(exc)})), 500

    @app.get("/api/project-portfolio/<slug>")
    def project_detail(slug: str) -> object:
        """Return a single project's detailed data."""
        try:
            data = load_project_performance_data()
            if "error" in data:
                return jsonify(json_safe({"ok": False, "error": data["error"]}))
            project = data.get("projects", {}).get(slug)
            if not project:
                return jsonify(json_safe({"ok": False, "error": f"Project '{slug}' not found"})), 404
            return jsonify(json_safe({"ok": True, "project": project}))
        except Exception as exc:
            return jsonify(json_safe({"ok": False, "error": str(exc)})), 500

    @app.get("/api/project-portfolio/<slug>/chart-data")
    def project_chart_data(slug: str) -> object:
        """Return pre-computed chart data for a project."""
        try:
            data = load_project_performance_data()
            if "error" in data:
                return jsonify(json_safe({"ok": False, "error": data["error"]}))
            project = data.get("projects", {}).get(slug)
            if not project:
                return jsonify(json_safe({"ok": False, "error": f"Project '{slug}' not found"})), 404

            sec_a = project.get("section_a", {})
            budget_lines = sec_a.get("budget_lines", [])
            sec_b = project.get("section_b", {})
            indicators = sec_b.get("indicators", [])
            sec_c = project.get("section_c", {})

            # ── Budget Burn Rate (bar: actual vs planned per budget line) ──
            burn_chart = {
                "categories": [bl["budget_line"] for bl in budget_lines],
                "planned": [bl["planned_cumulative"] for bl in budget_lines],
                "actual": [bl["actual_cumulative"] for bl in budget_lines],
            }

            # ── Indicator Achievement ──
            indicator_chart = {
                "categories": [ind["indicator"] for ind in indicators],
                "annual_targets": [ind["annual_target"] for ind in indicators],
                "actual_results": [ind["actual_cumulative"] for ind in indicators],
                "achievement_pcts": [ind["achievement_pct"] for ind in indicators],
            }

            # ── Budget Line RAG Distribution ──
            rag_dist = {"On Track": 0, "Watch": 0, "Off Track": 0, "N/A": 0}
            for bl in budget_lines:
                r = bl.get("rag", "N/A")
                if r in rag_dist:
                    rag_dist[r] += 1
                else:
                    rag_dist["N/A"] += 1

            return jsonify(json_safe({
                "ok": True,
                "slug": slug,
                "burn_chart": burn_chart,
                "indicator_chart": indicator_chart,
                "rag_distribution": rag_dist,
                "overall_rag": sec_c.get("overall_rag", "N/A"),
                "financial_rag": sec_c.get("financial_rag", "N/A"),
                "technical_rag": sec_c.get("technical_rag", "N/A"),
                "total_annual_budget": sec_a.get("total_annual_budget", 0),
                "total_expenditure": sec_a.get("total_cumulative_expenditure", 0),
                "total_variance_pct": sec_a.get("total_variance_pct"),
            }))
        except Exception as exc:
            return jsonify(json_safe({"ok": False, "error": str(exc)})), 500

    return app


def load_source_dataframe(csv_path: Path) -> pd.DataFrame:
    if not csv_path.exists():
        raise FileNotFoundError(
            f"Required source file not found: {csv_path.name}. The server must load this CSV at startup."
        )

    dataframe = pd.read_csv(csv_path)
    if dataframe.empty:
        raise ValueError("golden_executive_record.csv is empty.")
    return dataframe


def initialize_database(csv_path: Path) -> sqlite3.Connection:
    dataframe = load_source_dataframe(csv_path)
    connection = sqlite3.connect(":memory:", check_same_thread=False)
    connection.row_factory = sqlite3.Row
    dataframe.to_sql(TABLE_NAME, connection, index=False, if_exists="replace")
    return connection


def build_gemini_providers() -> list[dict[str, str]]:
    default_model = os.getenv("GEMINI_MODEL", "gemini-1.5-flash")
    provider_slots = [
        ("default", os.getenv("GEMINI_API_KEY"), os.getenv("GEMINI_MODEL", default_model)),
        ("1", os.getenv("GEMINI_API_KEY1"), os.getenv("GEMINI_MODEL1", default_model)),
        ("2", os.getenv("GEMINI_API_KEY2"), os.getenv("GEMINI_MODEL2", default_model)),
        ("3", os.getenv("GEMINI_API_KEY3"), os.getenv("GEMINI_MODEL3", default_model)),
        ("4", os.getenv("GEMINI_API_KEY4"), os.getenv("GEMINI_MODEL4", default_model)),
    ]

    providers: list[dict[str, str]] = []
    seen_keys: set[str] = set()
    for slot, api_key, model_name in provider_slots:
        if api_key and api_key not in seen_keys:
            providers.append({"slot": slot, "api_key": api_key, "model_name": model_name})
            seen_keys.add(api_key)
    return providers


def build_ai_system_instruction(schema_sql: str) -> str:
    return (
        "You are a precise text-to-SQL engine for a healthcare analytics dashboard. "
        "You must answer by generating only one SQLite SELECT statement. "
        "Do not explain your reasoning. Do not mention policies. Do not use markdown. "
        "Do not invent tables, columns, or external facts. Use only the schema provided below. "
        "If the user's question cannot be answered from this schema, return a single SELECT that yields a short error message as a column named message.\n\n"
        f"Schema:\n{schema_sql}\n\n"
        f"Allowed table: {TABLE_NAME}."
    )


def build_ai_prompt(question: str, chart_id: str = "") -> str:
    chart_context = f"\nCurrent chart context: {chart_id}" if chart_id else ""
    return (
        "Generate exactly one SQLite query for the question below. "
        "Return SQL only. No markdown fences, no commentary, no prose. "
        "Use only SELECT or WITH clauses. Never use INSERT, UPDATE, DELETE, DROP, ALTER, CREATE, PRAGMA, ATTACH, or multiple statements.\n\n"
        f"Question: {question}{chart_context}"
    )


def select_ai_provider(providers: list[dict[str, str]], router_state: dict[str, Any]) -> tuple[int | None, dict[str, str] | None]:
    if not providers:
        return None, None

    now = time.monotonic()
    start_index = int(router_state.get("cursor", 0)) % len(providers)
    cooldowns = router_state.setdefault("cooldowns", {})

    for offset in range(len(providers)):
        provider_index = (start_index + offset) % len(providers)
        if cooldowns.get(provider_index, 0) <= now:
            router_state["cursor"] = provider_index + 1
            return provider_index, providers[provider_index]

    return None, None


def mark_provider_cooldown(router_state: dict[str, Any], provider_index: int | None, seconds: int) -> None:
    if provider_index is None:
        return
    cooldowns = router_state.setdefault("cooldowns", {})
    cooldowns[provider_index] = time.monotonic() + seconds


def is_quota_error(exc: Exception) -> bool:
    message = str(exc).lower()
    return any(token in message for token in ["quota", "rate limit", "429", "exceeded"])


def json_safe(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): json_safe(item) for key, item in value.items()}

    if isinstance(value, list):
        return [json_safe(item) for item in value]

    if isinstance(value, tuple):
        return [json_safe(item) for item in value]

    if isinstance(value, pd.DataFrame):
        return json_safe(value.to_dict(orient="records"))

    if isinstance(value, pd.Series):
        return json_safe(value.tolist())

    try:
        if pd.isna(value):
            return None
    except Exception:
        pass

    if hasattr(value, "item") and not isinstance(value, (str, bytes)):
        try:
            return json_safe(value.item())
        except Exception:
            pass

    return value


def load_datim_location_hierarchy(workbook_path: Path) -> dict[str, Any]:
    hierarchy: dict[str, Any] = {
        "counties": [],
        "subcounties_by_county": {},
        "facilities_by_county": {},
        "facilities_by_subcounty": {},
        "subcounties": [],
        "facilities": [],
    }

    if not workbook_path.exists():
        return hierarchy

    ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

    def resolve_shared_items(field: ET.Element) -> list[str]:
        shared_items = field.find("main:sharedItems", ns)
        if shared_items is None:
            return []

        items: list[str] = []
        for child in list(shared_items):
            tag = child.tag.split("}")[-1]
            if tag == "s":
                items.append(clean_text(child.attrib.get("v")))
            elif tag == "n":
                items.append(clean_text(child.attrib.get("v")))
            elif tag == "m":
                items.append("")
            else:
                items.append(clean_text(child.attrib.get("v")))
        return items

    def resolve_record_value(cell: ET.Element, items: list[str]) -> str:
        tag = cell.tag.split("}")[-1]
        if tag == "x":
            try:
                index = int(cell.attrib.get("v", "0"))
            except ValueError:
                return ""
            return items[index] if 0 <= index < len(items) else ""
        if tag == "s":
            return clean_text(cell.attrib.get("v"))
        if tag == "n":
            return clean_text(cell.attrib.get("v"))
        return clean_text(cell.attrib.get("v"))

    try:
        with zipfile.ZipFile(workbook_path) as workbook:
            definition = ET.fromstring(workbook.read("xl/pivotCache/pivotCacheDefinition1.xml"))
            cache_fields = definition.find("main:cacheFields", ns)
            if cache_fields is None:
                return hierarchy

            fields = cache_fields.findall("main:cacheField", ns)
            field_names = [clean_text(field.attrib.get("name")) for field in fields]
            county_index = next((index for index, name in enumerate(field_names) if name.lower() == "county"), None)
            subcounty_index = next(
                (index for index, name in enumerate(field_names) if name.lower() in {"subcounty", "sub county", "sub-county"}),
                None,
            )

            if county_index is None or subcounty_index is None:
                return hierarchy

            county_items = resolve_shared_items(fields[county_index])
            subcounty_items = resolve_shared_items(fields[subcounty_index])
            facility_index = next((index for index, name in enumerate(field_names) if name.lower() in {"faility", "facility"}), None)
            facility_items = resolve_shared_items(fields[facility_index]) if facility_index is not None else []
            records_root = ET.fromstring(workbook.read("xl/pivotCache/pivotCacheRecords1.xml"))

            subcounty_mapping: dict[str, set[str]] = {}
            county_facility_mapping: dict[str, set[str]] = {}
            subcounty_facility_mapping: dict[str, dict[str, set[str]]] = {}
            all_facilities: set[str] = set()
            for record in records_root.findall("main:r", ns):
                cells = list(record)
                if len(cells) <= max(county_index, subcounty_index, facility_index or 0):
                    continue

                county = clean_text(resolve_record_value(cells[county_index], county_items))
                subcounty = clean_text(resolve_record_value(cells[subcounty_index], subcounty_items))
                facility = clean_text(resolve_record_value(cells[facility_index], facility_items)) if facility_index is not None else ""
                if not county or not subcounty:
                    continue

                subcounty_mapping.setdefault(county, set()).add(subcounty)
                if facility:
                    county_facility_mapping.setdefault(county, set()).add(facility)
                    subcounty_facility_mapping.setdefault(county, {}).setdefault(subcounty, set()).add(facility)
                    all_facilities.add(facility)

            hierarchy["counties"] = sorted(subcounty_mapping.keys())
            hierarchy["subcounties_by_county"] = {county: sorted(subcounties) for county, subcounties in subcounty_mapping.items()}
            hierarchy["facilities_by_county"] = {county: sorted(facilities) for county, facilities in county_facility_mapping.items()}
            hierarchy["facilities_by_subcounty"] = {
                county: {subcounty: sorted(facilities) for subcounty, facilities in sub_map.items()}
                for county, sub_map in subcounty_facility_mapping.items()
            }
            hierarchy["subcounties"] = sorted({subcounty for subcounties in subcounty_mapping.values() for subcounty in subcounties})
            hierarchy["facilities"] = sorted(all_facilities)
            return hierarchy
    except Exception:
        return hierarchy


def load_location_hierarchy_from_csv(csv_path: Path) -> dict[str, Any]:
    """Load location hierarchy from location_hierarchy.csv (County,SubCounty,Facility).
    Returns same dict structure as load_datim_location_hierarchy()."""
    import pandas as pd

    hierarchy: dict[str, Any] = {
        "counties": [],
        "subcounties_by_county": {},
        "facilities_by_county": {},
        "facilities_by_subcounty": {},
        "subcounties": [],
        "facilities": [],
    }

    if not csv_path.exists():
        return hierarchy

    try:
        df = pd.read_csv(csv_path)
        required = {"County", "SubCounty", "Facility"}
        if not required.issubset(df.columns):
            return hierarchy

        df = df.dropna(subset=["County", "SubCounty", "Facility"])

        subcounty_mapping: dict[str, set[str]] = {}
        county_facility_mapping: dict[str, set[str]] = {}
        subcounty_facility_mapping: dict[str, dict[str, set[str]]] = {}
        all_facilities: set[str] = set()

        for _, row in df.iterrows():
            county = str(row["County"]).strip()
            subcounty = str(row["SubCounty"]).strip()
            facility = str(row["Facility"]).strip()
            if not county or not subcounty or not facility:
                continue

            subcounty_mapping.setdefault(county, set()).add(subcounty)
            county_facility_mapping.setdefault(county, set()).add(facility)
            subcounty_facility_mapping.setdefault(county, {}).setdefault(subcounty, set()).add(facility)
            all_facilities.add(facility)

        hierarchy["counties"] = sorted(subcounty_mapping.keys())
        hierarchy["subcounties_by_county"] = {c: sorted(s) for c, s in subcounty_mapping.items()}
        hierarchy["facilities_by_county"] = {c: sorted(f) for c, f in county_facility_mapping.items()}
        hierarchy["facilities_by_subcounty"] = {
            c: {sc: sorted(f) for sc, f in sc_map.items()}
            for c, sc_map in subcounty_facility_mapping.items()
        }
        hierarchy["subcounties"] = sorted({sc for subs in subcounty_mapping.values() for sc in subs})
        hierarchy["facilities"] = sorted(all_facilities)
        return hierarchy
    except Exception:
        return hierarchy


def load_datim_hiv_treatment_sections(workbook_path: Path) -> list[dict[str, Any]]:
    sections = [
        {
            "label": "Newly Started on ART",
            "keywords": ["newly started on art", "new started on art", "sum of tbart new", "start on art"],
        },
        {
            "label": "Current on ART",
            "keywords": ["current on art", "previously enrolled on art", "enrolled on art"],
        },
        {
            "label": "ART Optimization",
            "keywords": ["art optimization", "tbart", "pmtct art"],
        },
        {
            "label": "Adverse Events - AE",
            "keywords": ["adverse events", "ae"],
        },
        {
            "label": "DSD",
            "keywords": ["dsd: hts_index", "dsd: tx_ml", "dsd: tx_rtt", "dsd"],
        },
        {
            "label": "VL Monitoring",
            "keywords": ["vl monitoring", "tx_pvls", "pvls", "routine"],
        },
        {
            "label": "Treatment Outcomes",
            "keywords": ["treatment outcomes", "iit", "rtt", "outcome"],
        },
        {
            "label": "OTZ",
            "keywords": ["otz"],
        },
        {
            "label": "OVC",
            "keywords": ["ovc"],
        },
        {
            "label": "COVID-19",
            "keywords": ["covid"],
        },
        {
            "label": "AHD",
            "keywords": ["ahd", "advanced hiv disease"],
        },
    ]

    if not workbook_path.exists():
        return [{"label": section["label"], "items": []} for section in sections]

    ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

    def load_shared_strings(workbook: zipfile.ZipFile) -> list[str]:
        if "xl/sharedStrings.xml" not in workbook.namelist():
            return []
        shared_root = ET.fromstring(workbook.read("xl/sharedStrings.xml"))
        values: list[str] = []
        seen: set[str] = set()
        for shared_item in shared_root.findall("main:si", ns):
            text_parts = [node.text or "" for node in shared_item.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t")]
            text = clean_text("".join(text_parts))
            if not text:
                continue
            key = text.lower()
            if key in seen:
                continue
            seen.add(key)
            values.append(text)
        return values

    try:
        with zipfile.ZipFile(workbook_path) as workbook:
            shared_strings = load_shared_strings(workbook)
    except Exception:
        shared_strings = []

    results: list[dict[str, Any]] = []
    for section in sections:
        label = section["label"]
        keywords = [keyword.lower() for keyword in section["keywords"]]
        matches: list[str] = []
        for text in shared_strings:
            lower_text = text.lower()
            if any(keyword in lower_text for keyword in keywords):
                matches.append(text)
        results.append(
            {
                "label": label,
                "items": matches[:8],
                "match_count": int(len(matches)),
            }
        )

    return results


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()


def to_int_safe(value: Any) -> int | None:
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass

    try:
        return int(float(str(value).strip()))
    except Exception:
        return None


def resolve_unit_context(
    unit_id: Any,
    org_units_map: dict[str, dict[str, Any]],
    hospital_map: dict[str, dict[str, Any]],
) -> dict[str, str]:
    unit_key = clean_text(unit_id)
    record = hospital_map.get(unit_key) or org_units_map.get(unit_key) or {}
    path = clean_text(record.get("path"))
    hierarchy: dict[int, str] = {}

    for ancestor_id in [segment for segment in path.split("/") if segment]:
        ancestor = hospital_map.get(ancestor_id) or org_units_map.get(ancestor_id)
        if not ancestor:
            continue
        level = to_int_safe(ancestor.get("level"))
        ancestor_name = clean_text(ancestor.get("hospital_name") or ancestor.get("name")) or ancestor_id
        if level is not None and level not in hierarchy:
            hierarchy[level] = ancestor_name

    facility_name = clean_text(record.get("hospital_name") or record.get("name")) or unit_key
    if 5 in hierarchy:
        facility_name = hierarchy[5]

    return {
        "Facility": facility_name,
        "County": hierarchy.get(3, ""),
        "SubCounty": hierarchy.get(4, ""),
    }


def normalize_dhis_analytics_frame(
    frame: pd.DataFrame,
    source_file: str,
    data_elements_map: dict[str, str],
    org_units_map: dict[str, dict[str, Any]],
    hospital_map: dict[str, dict[str, Any]],
) -> pd.DataFrame:
    normalized = frame.rename(columns={column: column.lower() for column in frame.columns}).copy()
    numeric_value = pd.to_numeric(normalized.get("value"), errors="coerce")

    transformed = pd.DataFrame()
    transformed["Indicator"] = normalized["dx"].astype(str).map(lambda key: clean_text(data_elements_map.get(key, key)) or key)
    transformed["Month"] = normalized["pe"].astype(str)

    facility_context = normalized["ou"].astype(str).map(lambda unit_id: resolve_unit_context(unit_id, org_units_map, hospital_map))
    transformed["Facility"] = facility_context.map(lambda context: context["Facility"])
    transformed["County"] = facility_context.map(lambda context: context["County"])
    transformed["SubCounty"] = facility_context.map(lambda context: context["SubCounty"])

    transformed["source_file"] = source_file
    transformed["Value"] = numeric_value

    lower_name = source_file.lower()
    if "finance" in lower_name or "spend" in lower_name:
        transformed["Monthly_Expenditure_USD"] = numeric_value
        transformed["Total_Visits"] = 0
    else:
        transformed["Total_Visits"] = numeric_value
        transformed["Monthly_Expenditure_USD"] = 0

    transformed["Cost_Per_ANC_Visit"] = 0
    return transformed


def format_month_label(month_value: Any) -> str:
    if pd.isna(month_value):
        return ""

    text = str(month_value).strip()
    if not text:
        return ""

    if text.isdigit():
        if len(text) == 6:
            return f"{text[:4]}-{text[4:]}"
        return text

    digits_only = re.sub(r"[^0-9]", "", text)
    if len(digits_only) == 6:
        return f"{digits_only[:4]}-{digits_only[4:]}"

    return text


def aggregate_facility_metrics(dataframe: pd.DataFrame) -> pd.DataFrame:
    aggregated = (
        dataframe.groupby("Facility", as_index=False)
        .agg(
            Total_Visits=("Total_Visits", "sum"),
            Monthly_Expenditure_USD=("Monthly_Expenditure_USD", "sum"),
            Avg_Cost_Per_ANC_Visit=("Cost_Per_ANC_Visit", "mean"),
        )
        .sort_values("Monthly_Expenditure_USD", ascending=False)
    )
    return aggregated



def build_canonical_catalog(dataframe: pd.DataFrame) -> dict[str, Any]:
    geography_available = any(
        column.lower() in {"geography", "county", "countyname", "county_name", "district", "region", "subcounty", "subcountyname", "sub_county", "sub_county_name"}
        for column in dataframe.columns
    )

    return {
        "source": CSV_PATH.name,
        "table": TABLE_NAME,
        "row_count": int(len(dataframe)),
        "columns": list(dataframe.columns),
        "dimensions": [
            {"key": "Month", "label": "Month", "type": "time", "drilldown": True},
            {"key": "Facility", "label": "Facility", "type": "organization", "drilldown": True},
            {"key": "County", "label": "County", "type": "geo", "drilldown": geography_available, "available": geography_available},
            {"key": "SubCounty", "label": "Sub-County", "type": "geo", "drilldown": geography_available, "available": geography_available},
            {"key": "Indicator", "label": "Indicator", "type": "indicator", "drilldown": True},
            {"key": "Geography", "label": "Geography", "type": "geo", "drilldown": geography_available, "available": geography_available},
        ],
        "indicators": [
            {"key": "Total_Visits", "label": "Total Visits", "aggregation": "sum", "measure": "volume"},
            {"key": "Monthly_Expenditure_USD", "label": "Monthly Expenditure USD", "aggregation": "sum", "measure": "finance"},
            {"key": "Cost_Per_ANC_Visit", "label": "Cost Per ANC Visit", "aggregation": "avg", "measure": "efficiency"},
        ],
        "drilldown_keys": ["Month", "County", "SubCounty", "Facility", "Indicator"],
        "derived_metrics": [
            {"key": "visits_per_usd", "label": "Visits per USD", "formula": "SUM(Total_Visits) / SUM(Monthly_Expenditure_USD)"},
            {"key": "avg_cost_per_visit", "label": "Average Cost Per Visit", "formula": "SUM(Monthly_Expenditure_USD) / SUM(Total_Visits)"},
            {"key": "monthly_visit_growth", "label": "Monthly Visit Growth", "formula": "Current month total visits - previous month total visits"},
        ],
        "notes": [
            "DHIS analytics rows are enriched with facility names plus county and sub-county labels from organisation units before being merged into the live dashboard dataset.",
            "Hospitals.csv is used as a lookup table for human-readable facility names and is not merged as raw fact rows.",
            "All numeric summaries, rankings, and grouped rollups are computed locally before any AI call.",
        ],
    }



def build_facility_page(
    dataframe: pd.DataFrame,
    page: int = 1,
    page_size: int = 20,
    search: str = "",
    location_hierarchy: dict[str, Any] | None = None,
    hospitals_frame: pd.DataFrame | None = None,
    org_units_map: dict[str, dict[str, Any]] | None = None,
    hospital_map: dict[str, dict[str, Any]] | None = None,
) -> dict[str, Any]:
    if location_hierarchy and location_hierarchy.get("facilities_by_subcounty"):
        rows: list[dict[str, Any]] = []
        facilities_by_subcounty = location_hierarchy.get("facilities_by_subcounty", {}) or {}
        for county, sub_map in facilities_by_subcounty.items():
            for subcounty, facilities in (sub_map or {}).items():
                for facility in facilities or []:
                    rows.append(
                        {
                            "Facility": facility,
                            "County": county,
                            "SubCounty": subcounty,
                            "Facility_ID": "",
                        }
                    )

        facility_frame = pd.DataFrame(rows)
        if search:
            facility_frame = facility_frame[facility_frame["Facility"].str.lower().str.contains(search, na=False)]

        total_rows = int(len(facility_frame))
        start = (page - 1) * page_size
        end = start + page_size
        page_frame = facility_frame.iloc[start:end].copy()

        return {
            "page": page,
            "page_size": page_size,
            "total_rows": total_rows,
            "rows": page_frame.to_dict(orient="records"),
            "has_more": end < total_rows,
        }

    if hospitals_frame is not None and not hospitals_frame.empty:
        facility_frame = hospitals_frame.copy()
        facility_frame["Facility"] = facility_frame.get("hospital_name", "").map(clean_text)
        facility_frame["Facility_ID"] = facility_frame.get("hospital_id", "").map(clean_text)

        org_units_map = org_units_map or {}
        hospital_map = hospital_map or {}
        if "hospital_id" in facility_frame.columns:
            context = facility_frame["hospital_id"].map(lambda unit_id: resolve_unit_context(unit_id, org_units_map, hospital_map))
            if "County" not in facility_frame.columns:
                facility_frame["County"] = context.map(lambda item: item["County"])
            if "SubCounty" not in facility_frame.columns:
                facility_frame["SubCounty"] = context.map(lambda item: item["SubCounty"])
            facility_frame["Facility"] = facility_frame["Facility"].where(facility_frame["Facility"].astype(str).str.strip() != "", context.map(lambda item: item["Facility"]))

        if search:
            facility_frame = facility_frame[facility_frame["Facility"].str.lower().str.contains(search, na=False)]

        total_rows = int(len(facility_frame))
        start = (page - 1) * page_size
        end = start + page_size
        page_frame = facility_frame.iloc[start:end].copy()

        return {
            "page": page,
            "page_size": page_size,
            "total_rows": total_rows,
            "rows": page_frame.to_dict(orient="records"),
            "has_more": end < total_rows,
        }

    facility_totals = aggregate_facility_metrics(dataframe).copy()
    if search:
        facility_totals = facility_totals[facility_totals["Facility"].str.lower().str.contains(search, na=False)]

    total_rows = int(len(facility_totals))
    start = (page - 1) * page_size
    end = start + page_size
    page_frame = facility_totals.iloc[start:end].copy()

    return {
        "page": page,
        "page_size": page_size,
        "total_rows": total_rows,
        "rows": page_frame.to_dict(orient="records"),
        "has_more": end < total_rows,
    }




def build_local_sql(question: str) -> str | None:
    normalized = re.sub(r"\s+", " ", question.lower()).strip()
    if not normalized:
        return None

    limit_match = re.search(r"\b(top|highest|first)\s+(\d+)\b", normalized)
    if limit_match:
        limit = int(limit_match.group(2))
    else:
        quantity_match = re.search(r"\b(\d+)\b", normalized)
        limit = int(quantity_match.group(1)) if quantity_match else 5

    spend_terms = ["monthly_expenditure_usd", "monthly expenditure", "budget", "cost", "spend"]
    visit_terms = ["total_visits", "visits", "volume", "attendance", "patients"]

    facility_scope_cte = (
        "WITH facility_scope AS ( "
        "SELECT Facility, Total_Visits, Monthly_Expenditure_USD, Cost_Per_ANC_Visit "
        "FROM clinics "
        "WHERE TRIM(COALESCE(Facility, '')) <> '' "
        "AND LOWER(TRIM(COALESCE(Facility, ''))) NOT IN ('kenya', 'unknown facility', 'national') "
        ") "
    )

    def ranked_facility_sql(order_column: str, order_direction: str, limit: int) -> str:
        return (
            facility_scope_cte
            + "SELECT Facility, "
            + "SUM(Total_Visits) AS Total_Visits, "
            + "ROUND(SUM(Monthly_Expenditure_USD), 2) AS Monthly_Expenditure_USD, "
            + "ROUND(AVG(Cost_Per_ANC_Visit), 2) AS Cost_Per_ANC_Visit "
            + "FROM facility_scope GROUP BY Facility "
            + f"ORDER BY {order_column} {order_direction} LIMIT {limit}"
        )

    if "month" in normalized or "trend" in normalized or "over time" in normalized:
        return (
            "SELECT Month, SUM(Total_Visits) AS total_visits, "
            "ROUND(SUM(Monthly_Expenditure_USD), 2) AS monthly_expenditure_usd, "
            "ROUND(AVG(Cost_Per_ANC_Visit), 2) AS avg_cost_per_anc_visit "
            "FROM clinics GROUP BY Month ORDER BY Month"
        )

    if any(token in normalized for token in spend_terms):
        order_direction = "ASC" if any(token in normalized for token in ["lowest", "least", "smallest", "bottom"]) else "DESC"
        return ranked_facility_sql("Monthly_Expenditure_USD", order_direction, limit)

    if any(token in normalized for token in visit_terms):
        order_direction = "ASC" if any(token in normalized for token in ["lowest", "least", "smallest", "bottom"]) else "DESC"
        return ranked_facility_sql("Total_Visits", order_direction, limit)

    if "average" in normalized or "avg" in normalized:
        if any(token in normalized for token in spend_terms):
            return "SELECT ROUND(AVG(Monthly_Expenditure_USD), 2) AS average_monthly_expenditure_usd FROM clinics"
        if any(token in normalized for token in visit_terms):
            return "SELECT ROUND(AVG(Total_Visits), 2) AS average_total_visits FROM clinics"

    if "total" in normalized or "sum" in normalized:
        if any(token in normalized for token in spend_terms):
            return "SELECT ROUND(SUM(Monthly_Expenditure_USD), 2) AS total_monthly_expenditure_usd FROM clinics"
        if any(token in normalized for token in visit_terms):
            return "SELECT SUM(Total_Visits) AS total_visits FROM clinics"

    if "clinic" in normalized or "facility" in normalized:
        return ranked_facility_sql("Monthly_Expenditure_USD", "DESC", limit)

    return None


def create_gemini_model(schema_sql: str):
    api_key = os.getenv("GEMINI_API_KEY")
    model_name = os.getenv("GEMINI_MODEL", "gemini-1.5-flash")

    if not api_key or genai is None:
        return None

    genai.configure(api_key=api_key)
    system_instruction = (
        "You are a precise text-to-SQL engine for a healthcare analytics dashboard. "
        "You must answer by generating only one SQLite SELECT statement. "
        "Do not explain your reasoning. Do not mention policies. Do not use markdown. "
        "Do not invent tables, columns, or external facts. Use only the schema provided below. "
        "If the user's question cannot be answered from this schema, return a single SELECT that yields a short error message as a column named message.\n\n"
        f"Schema:\n{schema_sql}\n\n"
        f"Allowed table: {TABLE_NAME}."
    )
    return genai.GenerativeModel(model_name=model_name, system_instruction=system_instruction)


def generate_sql(
    providers: list[dict[str, str]],
    router_state: dict[str, Any],
    question: str,
    allowed_columns: list[str],
    schema_sql: str,
    chart_id: str = "",
) -> tuple[str, str]:
    # AI-only mode: require a working Gemini provider and do not fall back to local heuristics.
    if not providers:
        raise ValueError("No Gemini providers are configured. The chat service requires at least one provider.")

    prompt = build_ai_prompt(question, chart_id)
    system_instruction = build_ai_system_instruction(schema_sql)
    last_error: str | None = None

    for _ in range(max(1, len(providers))):
        provider_index, provider = select_ai_provider(providers, router_state)
        if provider is None:
            break

        try:
            if genai is None:
                raise RuntimeError("Gemini SDK is unavailable.")

            genai.configure(api_key=provider["api_key"])
            model = genai.GenerativeModel(
                model_name=provider["model_name"],
                system_instruction=system_instruction,
            )
            response = model.generate_content(prompt)
            raw_text = getattr(response, "text", "") or ""
            sql_text = extract_sql(raw_text)
            if not sql_text:
                raise ValueError("Gemini did not return a SQL query.")
            return sql_text, f"gemini:{provider['slot']}", None
        except Exception as exc:
            last_error = str(exc)
            cooldown_seconds = 300 if is_quota_error(exc) else 45
            mark_provider_cooldown(router_state, provider_index, cooldown_seconds)
            continue

    # If we reach here no provider returned a SQL query.
    return None, "ai_error", last_error


def extract_sql(text: str) -> str:
    cleaned_text = text.strip()
    fenced_match = re.search(r"```sql\s*(.*?)\s*```", cleaned_text, flags=re.IGNORECASE | re.DOTALL)
    if fenced_match:
        cleaned_text = fenced_match.group(1).strip()
    cleaned_text = re.sub(r"^```(?:sql)?|```$", "", cleaned_text, flags=re.IGNORECASE).strip()
    return cleaned_text


def build_fallback_sql(question: str, allowed_columns: list[str]) -> str:
    normalized = re.sub(r"\s+", " ", question.lower()).strip()
    limit_match = re.search(r"\b(top|highest|first)\s+(\d+)\b", normalized)
    if limit_match:
        limit = int(limit_match.group(2))
    else:
        quantity_match = re.search(r"\b(\d+)\b", normalized)
        limit = int(quantity_match.group(1)) if quantity_match else 5

    spend_columns = ["monthly_expenditure_usd", "monthly expenditure", "budget", "cost", "spend"]
    visit_columns = ["total_visits", "visits", "volume", "attendance", "patients"]

    if any(token in normalized for token in ["how many", "count", "number of"]):
        if any(token in normalized for token in ["hospital", "hospitals", "facility", "facilities", "clinic", "clinics"]):
            return (
                "SELECT COUNT(DISTINCT Facility) AS hospital_count "
                "FROM clinics "
                "WHERE TRIM(COALESCE(Facility, '')) <> '' "
                "AND LOWER(TRIM(COALESCE(Facility, ''))) NOT IN ('kenya', 'unknown facility', 'national')"
            )
        if any(token in normalized for token in ["county", "counties"]):
            return (
                "SELECT COUNT(DISTINCT County) AS county_count "
                "FROM clinics "
                "WHERE TRIM(COALESCE(County, '')) <> ''"
            )
        if any(token in normalized for token in ["sub-county", "subcounty", "sub counties", "subcounties"]):
            return (
                "SELECT COUNT(DISTINCT SubCounty) AS subcounty_count "
                "FROM clinics "
                "WHERE TRIM(COALESCE(SubCounty, '')) <> ''"
            )

    if any(token in normalized for token in spend_columns):
        order_direction = "ASC" if any(token in normalized for token in ["lowest", "least", "smallest", "bottom"]) else "DESC"
        return (
            "SELECT Facility, Total_Visits, Monthly_Expenditure_USD, Cost_Per_ANC_Visit "
            f"FROM clinics ORDER BY Monthly_Expenditure_USD {order_direction} LIMIT {limit}"
        )

    if any(token in normalized for token in visit_columns):
        order_direction = "ASC" if any(token in normalized for token in ["lowest", "least", "smallest", "bottom"]) else "DESC"
        return (
            "SELECT Facility, Total_Visits, Monthly_Expenditure_USD, Cost_Per_ANC_Visit "
            f"FROM clinics ORDER BY Total_Visits {order_direction} LIMIT {limit}"
        )

    if "average" in normalized or "avg" in normalized:
        if any(token in normalized for token in spend_columns):
            return "SELECT ROUND(AVG(Monthly_Expenditure_USD), 2) AS average_monthly_expenditure_usd FROM clinics"
        if any(token in normalized for token in visit_columns):
            return "SELECT ROUND(AVG(Total_Visits), 2) AS average_total_visits FROM clinics"

    if "total" in normalized or "sum" in normalized:
        if any(token in normalized for token in spend_columns):
            return "SELECT ROUND(SUM(Monthly_Expenditure_USD), 2) AS total_monthly_expenditure_usd FROM clinics"
        if any(token in normalized for token in visit_columns):
            return "SELECT SUM(Total_Visits) AS total_visits FROM clinics"

    if "clinic" in normalized or "facility" in normalized:
        return (
            "SELECT Facility, Total_Visits, Monthly_Expenditure_USD, Cost_Per_ANC_Visit "
            f"FROM clinics ORDER BY Monthly_Expenditure_USD DESC LIMIT {limit}"
        )

    return "SELECT Facility, Total_Visits, Monthly_Expenditure_USD, Cost_Per_ANC_Visit FROM clinics ORDER BY Monthly_Expenditure_USD DESC LIMIT 5"


def validate_sql(sql_text: str, allowed_columns: list[str]) -> str:
    candidate = sql_text.strip()
    normalized = re.sub(r"\s+", " ", candidate).lower()

    if not candidate:
        raise ValueError("Empty SQL query returned by the model.")

    if ";" in candidate:
        raise ValueError("Multiple SQL statements are not allowed.")

    if not (normalized.startswith("select ") or normalized.startswith("with ")):
        raise ValueError("Only SELECT queries are allowed.")

    forbidden_terms = [
        " insert ",
        " update ",
        " delete ",
        " drop ",
        " alter ",
        " create ",
        " attach ",
        " pragma ",
        " vacuum ",
        " begin ",
        " commit ",
        " rollback ",
        " replace ",
        " detach ",
    ]
    padded = f" {normalized} "
    if any(term in padded for term in forbidden_terms):
        raise ValueError("Unsafe SQL keyword detected in the generated query.")

    if re.search(r"\bsqlite_master\b|\bsqlite_schema\b", normalized):
        raise ValueError("System tables are not allowed.")

    if not re.search(r"\bfrom\b|\bjoin\b", normalized):
        raise ValueError("SQL must query data from the clinics table or a CTE derived from it.")

    allowed_column_tokens = {f'"{column}"' for column in allowed_columns}
    allowed_column_tokens.update({f"[{column}]" for column in allowed_columns})
    _ = allowed_column_tokens  # Reserved for future stricter identifier validation.

    return candidate


def run_safe_query(connection: sqlite3.Connection, sql_text: str) -> pd.DataFrame:
    with connection:
        dataframe = pd.read_sql_query(sql_text, connection)

    if len(dataframe) > MAX_RESULT_ROWS:
        dataframe = dataframe.head(MAX_RESULT_ROWS).copy()
        dataframe.attrs["truncated"] = True
    else:
        dataframe.attrs["truncated"] = False

    return dataframe


def build_chat_response(question: str, sql_text: str, result_frame: pd.DataFrame, source: str, chart_id: str = "") -> dict:
    row_count = int(len(result_frame))
    truncated = bool(result_frame.attrs.get("truncated", False))
    normalized_question = question.lower()

    if result_frame.empty:
        if any(term in normalized_question for term in ["clinic", "clinics", "facility", "facilities"]) and any(
            term in normalized_question for term in ["spend", "expenditure", "budget", "cost"]
        ):
            summary = (
                "No clinic-level expenditure rows are available in the loaded CSVs. "
                "The current data is aggregated at county/national level, so clinic rankings cannot be produced reliably."
            )
        else:
            summary = f"No rows matched the question: {question}"
        html_block = f"<p class=\"text-sm text-slate-600\">{html.escape(summary)}</p>"
        return {
            "question": question,
            "sql": sql_text,
            "source": source,
            "summary": summary,
            "answer_html": html_block,
            "row_count": 0,
            "truncated": False,
            "columns": list(result_frame.columns),
            "rows": [],
        }

    summary = f"Returned {row_count} row{'s' if row_count != 1 else ''}."
    if source == "fallback":
        summary += " Gemini is rate-limited right now, so a local SQLite fallback answered this one."
    if source == "local":
        summary += " This answer was computed locally to avoid an unnecessary model call."
    if source == "no_gemini":
        summary += " Gemini is not configured in this environment, so a local SQLite fallback answered this one."
    if chart_id:
        summary += f" Chart context: {chart_id}."
    if truncated:
        summary += f" Showing the first {MAX_RESULT_ROWS}."

    html_block = [f"<p class=\"text-sm text-slate-600 mb-3\">{html.escape(summary)}</p>"]
    html_block.append(render_html_table(result_frame))

    return {
        "question": question,
        "sql": sql_text,
        "source": source,
        "summary": summary,
        "answer_html": "".join(html_block),
        "row_count": row_count,
        "truncated": truncated,
        "columns": list(result_frame.columns),
        "rows": result_frame.to_dict(orient="records"),
    }


def render_html_table(dataframe: pd.DataFrame) -> str:
    header_cells = "".join(
        f'<th class="px-3 py-2 text-left text-xs font-semibold uppercase tracking-wide text-slate-500 border-b border-slate-200">{html.escape(str(column))}</th>'
        for column in dataframe.columns
    )

    body_rows = []
    for row in dataframe.itertuples(index=False, name=None):
        cells = []
        for value in row:
            cell_value = "" if pd.isna(value) else str(value)
            cells.append(
                f'<td class="px-3 py-2 text-sm text-slate-700 border-b border-slate-100 whitespace-nowrap">{html.escape(cell_value)}</td>'
            )
        body_rows.append(f"<tr>{''.join(cells)}</tr>")

    return (
        '<div class="overflow-x-auto rounded-xl border border-slate-200 bg-white shadow-sm">'
        '<table class="min-w-full divide-y divide-slate-200">'
        f"<thead><tr>{header_cells}</tr></thead>"
        f"<tbody>{''.join(body_rows)}</tbody>"
        "</table></div>"
    )


# ────────────────────────────────────────────────────────────
# PROJECT PERFORMANCE MONITORING — Excel Data Loader
# ────────────────────────────────────────────────────────────
PERFORMANCE_XLSX_PATH = BASE_DIR.parent / "CHAK Monthly Project Performance Monitoring Template - Topline Indicators.xlsx"

_PROJECT_CACHE: dict[str, Any] | None = None
_PROJECT_CACHE_MTIME: float = 0

PROJECT_SHEET_MAP = {
    "jamii-tekelezi": "Jamii Tekelezi",
    "chap-stawisha": "CHAP Stawisha",
    "eye-health": "Eye Health - ACSP & GitLab",
    "eis": "EIS",
    "bftw-hss": "BFTW HSS",
    "bftw-rmncah": "BFTW RMNCAH",
    "pep": "PEP",
    "gf-mnch": "GF-MNCH",
    "impact": "IMPACT",
    "cdic-icare": "CDIC-iCARE",
}

BUDGET_LINE_NAMES = [
    "Personnel & HR",
    "Technical / Program Activities",
    "Equipment, Supplies & Commodities",
    "Training & Capacity Building",
    "Sub-awards / Partner Disbursements",
    "Monitoring, Evaluation & Learning",
    "Travel & Transport",
    "Administration & Overheads",
    "Other Direct Costs / Contingency",
]


def _parse_cell(cell: Any) -> float | str | None:
    """Parse a cell value: return float if numeric, stripped string if text, None if blank."""
    # Handle openpyxl Cell objects
    if hasattr(cell, "value"):
        value = cell.value
    else:
        value = cell
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, datetime):
        return value.strftime("%B %Y")
    s = str(value).strip()
    if not s:
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return s


def _rag_class(rag: str | None) -> str:
    """Normalize RAG string."""
    if not rag:
        return "N/A"
    r = str(rag).strip().lower()
    if r == "on track":
        return "On Track"
    if r == "watch":
        return "Watch"
    if r == "off track":
        return "Off Track"
    return str(rag).strip()


def load_project_performance_data(force_reload: bool = False) -> dict[str, Any]:
    """Load and cache the project performance Excel workbook. Returns a dict with
    'portfolio' (aggregate dashboard) and 'projects' (per-project detail)."""
    global _PROJECT_CACHE, _PROJECT_CACHE_MTIME

    xlsx = PERFORMANCE_XLSX_PATH
    if not xlsx.exists():
        return {"error": f"File not found: {xlsx.name}"}

    current_mtime = xlsx.stat().st_mtime
    if not force_reload and _PROJECT_CACHE is not None and current_mtime == _PROJECT_CACHE_MTIME:
        return _PROJECT_CACHE

    import openpyxl
    wb = openpyxl.load_workbook(xlsx, data_only=True)

    # ── Portfolio Dashboard Sheet ──
    portfolio_raw = _parse_portfolio_dashboard(wb["Portfolio Dashboard"])
    # ── Project Sheets ──
    projects: dict[str, Any] = {}
    for slug, sheet_name in PROJECT_SHEET_MAP.items():
        if sheet_name in wb.sheetnames:
            projects[slug] = _parse_project_sheet(wb[sheet_name], slug)

    wb.close()

    result = {
        "portfolio": portfolio_raw,
        "projects": projects,
        "loaded_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "sheet_count": len(PROJECT_SHEET_MAP),
    }
    _PROJECT_CACHE = result
    _PROJECT_CACHE_MTIME = current_mtime
    return result


def _parse_portfolio_dashboard(ws: Any) -> dict[str, Any]:
    """Parse the Portfolio Dashboard sheet."""
    rows_data: list[dict[str, Any]] = []
    # Data rows 6-15 (Excel rows); row 5 is header
    for excel_row in range(6, 16):
        proj = _parse_cell(ws.cell(row=excel_row, column=2))
        if not proj or proj == "PORTFOLIO TOTAL":
            break
        rows_data.append({
            "project": str(proj),
            "donor": str(_parse_cell(ws.cell(row=excel_row, column=3)) or ""),
            "total_annual_budget": _parse_cell(ws.cell(row=excel_row, column=4)) or 0,
            "cumulative_expenditure": _parse_cell(ws.cell(row=excel_row, column=5)) or 0,
            "budget_variance_pct": _parse_cell(ws.cell(row=excel_row, column=6)),
            "financial_rag": _rag_class(_parse_cell(ws.cell(row=excel_row, column=7))),
            "technical_rag": _rag_class(_parse_cell(ws.cell(row=excel_row, column=8))),
            "indicators_off_track": str(_parse_cell(ws.cell(row=excel_row, column=9)) or ""),
            "overall_rag": _rag_class(_parse_cell(ws.cell(row=excel_row, column=10))),
        })

    # Portfolio total (row 16)
    total_annual = _parse_cell(ws.cell(row=16, column=4)) or 0
    total_expenditure = _parse_cell(ws.cell(row=16, column=5)) or 0

    # CEO Snapshot (rows 19-20)
    on_track_count = str(_parse_cell(ws.cell(row=19, column=4)) or "")
    on_watch_count = str(_parse_cell(ws.cell(row=19, column=7)) or "")
    off_track_count = str(_parse_cell(ws.cell(row=20, column=4)) or "")
    budget_utilisation = _parse_cell(ws.cell(row=20, column=7))

    return {
        "projects": rows_data,
        "portfolio_total_annual_budget": total_annual,
        "portfolio_total_expenditure": total_expenditure,
        "ceo_snapshot": {
            "on_track": on_track_count,
            "on_watch": on_watch_count,
            "off_track": off_track_count,
            "budget_utilisation_pct": round(budget_utilisation * 100, 1) if isinstance(budget_utilisation, (int, float)) else None,
        },
    }


def _parse_project_sheet(ws: Any, slug: str) -> dict[str, Any]:
    """Parse an individual project sheet. Returns structured sections A, B, C."""
    # ── Header ──
    donor = str(_parse_cell(ws.cell(row=4, column=3)) or "")
    project_code = str(_parse_cell(ws.cell(row=5, column=3)) or "")
    reporting_month = ws.cell(row=4, column=7).value
    project_duration = _parse_cell(ws.cell(row=4, column=11))
    months_elapsed = _parse_cell(ws.cell(row=5, column=11))

    # Format reporting month
    if isinstance(reporting_month, datetime):
        reporting_month_str = reporting_month.strftime("%B %Y")
    else:
        reporting_month_str = str(reporting_month or "")

    # ── SECTION A: Financial Performance ──
    budget_lines: list[dict[str, Any]] = []
    for i, name in enumerate(BUDGET_LINE_NAMES):
        row = 9 + i
        budget_lines.append({
            "budget_line": name,
            "annual_budget": _parse_cell(ws.cell(row=row, column=3)) or 0,
            "planned_cumulative": _parse_cell(ws.cell(row=row, column=4)) or 0,
            "actual_cumulative": _parse_cell(ws.cell(row=row, column=5)) or 0,
            "variance_amount": _parse_cell(ws.cell(row=row, column=6)),
            "variance_pct": _parse_cell(ws.cell(row=row, column=7)),
            "current_month_expenditure": _parse_cell(ws.cell(row=row, column=8)) or 0,
            "avg_monthly_burn_rate": _parse_cell(ws.cell(row=row, column=9)) or 0,
            "projected_annual_expenditure": _parse_cell(ws.cell(row=row, column=10)) or 0,
            "months_remaining": _parse_cell(ws.cell(row=row, column=11)),
            "rag": _rag_class(_parse_cell(ws.cell(row=row, column=12))),
        })

    # Total row (row 18)
    total_budget_line = {
        "budget_line": "TOTAL PROJECT BUDGET",
        "annual_budget": _parse_cell(ws.cell(row=18, column=3)) or 0,
        "planned_cumulative": _parse_cell(ws.cell(row=18, column=4)) or 0,
        "actual_cumulative": _parse_cell(ws.cell(row=18, column=5)) or 0,
        "variance_amount": _parse_cell(ws.cell(row=18, column=6)),
        "variance_pct": _parse_cell(ws.cell(row=18, column=7)),
        "current_month_expenditure": _parse_cell(ws.cell(row=18, column=8)) or 0,
        "avg_monthly_burn_rate": _parse_cell(ws.cell(row=18, column=9)) or 0,
        "projected_annual_expenditure": _parse_cell(ws.cell(row=18, column=10)) or 0,
        "months_remaining": _parse_cell(ws.cell(row=18, column=11)),
        "rag": _rag_class(_parse_cell(ws.cell(row=18, column=12))),
    }

    # ── SECTION B: Technical / Donor Indicators ──
    indicators: list[dict[str, Any]] = []
    for row in range(23, 31):
        ind_name = _parse_cell(ws.cell(row=row, column=2))
        if not ind_name:
            continue
        indicators.append({
            "indicator": str(ind_name),
            "definition": str(_parse_cell(ws.cell(row=row, column=3)) or ""),
            "annual_target": _parse_cell(ws.cell(row=row, column=4)) or 0,
            "planned_cumulative": _parse_cell(ws.cell(row=row, column=5)) or 0,
            "actual_cumulative": _parse_cell(ws.cell(row=row, column=6)) or 0,
            "achievement_pct": _parse_cell(ws.cell(row=row, column=7)),
            "rag": _rag_class(_parse_cell(ws.cell(row=row, column=12))),
        })

    # ── SECTION C: Overall Project Health ──
    financial_rag = _rag_class(_parse_cell(ws.cell(row=34, column=5)))
    off_track_budget_lines = str(_parse_cell(ws.cell(row=34, column=10)) or "")
    technical_rag = _rag_class(_parse_cell(ws.cell(row=35, column=5)))
    off_track_indicators = str(_parse_cell(ws.cell(row=35, column=10)) or "")
    overall_rag = _rag_class(_parse_cell(ws.cell(row=36, column=5)))

    # Compute financial variance pct from total budget line
    total_annual = total_budget_line["annual_budget"] or 0
    total_actual = total_budget_line["actual_cumulative"] or 0
    total_planned = total_budget_line["planned_cumulative"] or 0
    total_variance_pct = total_budget_line["variance_pct"]

    return {
        "slug": slug,
        "donor": donor,
        "project_code": project_code,
        "reporting_month": reporting_month_str,
        "project_duration_months": project_duration,
        "months_elapsed": months_elapsed,
        "section_a": {
            "budget_lines": budget_lines,
            "total": total_budget_line,
            "total_annual_budget": total_annual,
            "total_cumulative_expenditure": total_actual,
            "total_planned_cumulative": total_planned,
            "total_variance_pct": total_variance_pct,
        },
        "section_b": {
            "indicators": indicators,
        },
        "section_c": {
            "financial_rag": financial_rag,
            "off_track_budget_lines": off_track_budget_lines,
            "technical_rag": technical_rag,
            "off_track_indicators": off_track_indicators,
            "overall_rag": overall_rag,
        },
    }


app = create_app()


if __name__ == "__main__":
    port = int(os.getenv("PORT", "5000"))
    # Debug: print registered routes
    rules = sorted([r.rule for r in app.url_map.iter_rules()])
    print(f"Registered {len(rules)} routes:")
    for r in rules:
        if "dhis" in r or "debug" in r:
            print(f"  *** {r}")
    app.run(host="127.0.0.1", port=port, debug=False, use_reloader=False, threaded=True)